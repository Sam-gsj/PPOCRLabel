# Do imports like python3 so our package works for 2 and 3
from __future__ import absolute_import


from tablepyxl.style import Table
from paddle.utils import try_import
from openpyxl import load_workbook
import openpyxl
from bs4 import BeautifulSoup
import os
from openpyxl.utils.exceptions import IllegalCharacterError


def string_to_int(s):
    if s.isdigit():
        return int(s)
    return 0


def get_Tables(doc):
    try_import("lxml")
    from lxml import etree, html

    tree = html.fromstring(doc)
    comments = tree.xpath("//comment()")
    for comment in comments:
        comment.drop_tag()
    return [Table(table) for table in tree.xpath("//table")]


def write_rows(worksheet, elem, row, column=1):
    """
    Writes every tr child element of elem to a row in the worksheet
    returns the next row after all rows are written
    """
    try_import("openpyxl")
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    initial_column = column
    for table_row in elem.rows:
        for table_cell in table_row.cells:
            cell = worksheet.cell(row=row, column=column)
            while isinstance(cell, MergedCell):
                column += 1
                cell = worksheet.cell(row=row, column=column)

            colspan = string_to_int(table_cell.element.get("colspan", "1"))
            rowspan = string_to_int(table_cell.element.get("rowspan", "1"))
            if rowspan > 1 or colspan > 1:
                worksheet.merge_cells(
                    start_row=row,
                    start_column=column,
                    end_row=row + rowspan - 1,
                    end_column=column + colspan - 1,
                )

            cell.value = table_cell.value
            table_cell.format(cell)
            min_width = table_cell.get_dimension("min-width")
            max_width = table_cell.get_dimension("max-width")

            if colspan == 1:
                # Initially, when iterating for the first time through the loop, the width of all the cells is None.
                # As we start filling in contents, the initial width of the cell (which can be retrieved by:
                # worksheet.column_dimensions[get_column_letter(column)].width) is equal to the width of the previous
                # cell in the same column (i.e. width of A2 = width of A1)
                width = max(
                    worksheet.column_dimensions[get_column_letter(column)].width or 0,
                    len(table_cell.value) + 2,
                )
                if max_width and width > max_width:
                    width = max_width
                elif min_width and width < min_width:
                    width = min_width
                worksheet.column_dimensions[get_column_letter(column)].width = width
            column += colspan
        row += 1
        column = initial_column
    return row


def table_to_sheet(table, wb):
    """
    Takes a table and workbook and writes the table to a new sheet.
    The sheet title will be the same as the table attribute name.
    """
    ws = wb.create_sheet(title=table.element.get("name"))
    insert_table(table, ws, 1, 1)


def document_to_workbook(doc, wb=None, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document.
    The workbook is returned
    """
    try_import("premailer")
    try_import("openpyxl")
    from premailer import Premailer
    from openpyxl import Workbook

    if not wb:
        wb = Workbook()
        wb.remove(wb.active)

    inline_styles_doc = Premailer(
        doc, base_url=base_url, remove_classes=False
    ).transform()
    tables = get_Tables(inline_styles_doc)

    for table in tables:
        table_to_sheet(table, wb)

    return wb


def document_to_xl(doc, filename, base_url=None):
    """
    Takes a string representation of an html document and writes one sheet for
    every table in the document. The workbook is written out to a file called filename
    """
    try:
        wb = document_to_workbook(doc, base_url=base_url)
        wb.save(filename)
    except Exception as e:
        print(f"保存 Excel 文件时出错: {e}")


def insert_table(table, worksheet, column, row):
    if table.head:
        row = write_rows(worksheet, table.head, row, column)
    if table.body:
        row = write_rows(worksheet, table.body, row, column)


def xl_to_html(excel_file):
    try:
        wb = load_workbook(excel_file)
        sheet_names = wb.sheetnames
        html_content = "<html><body>"

        for sheet_name in sheet_names:
            try:
                sheet = wb[sheet_name]
                html_content += f"<h2>{sheet_name}</h2>"
                html_content += "<table border='1'>"

                # 获取合并单元格信息
                merged_cells = sheet.merged_cells.ranges

                # 创建一个字典来存储合并单元格的范围
                merged_cell_map = {}
                for merged_cell in merged_cells:
                    for row in range(merged_cell.min_row, merged_cell.max_row + 1):
                        for col in range(merged_cell.min_col, merged_cell.max_col + 1):
                            merged_cell_map[(row, col)] = {
                                "range": merged_cell,
                                "main_cell": (merged_cell.min_row, merged_cell.min_col),
                            }

                # 遍历每一行
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
                    html_content += "<tr>"

                    # 遍历每一列
                    for col_idx, cell_value in enumerate(row, 1):
                        current_pos = (row_idx, col_idx)

                        # 检查当前单元格是否在合并单元格范围内
                        if current_pos in merged_cell_map:
                            merge_info = merged_cell_map[current_pos]
                            merge_range = merge_info["range"]

                            # 只有主单元格（左上角单元格）需要输出
                            if current_pos == merge_info["main_cell"]:
                                rowspan = merge_range.max_row - merge_range.min_row + 1
                                colspan = merge_range.max_col - merge_range.min_col + 1
                                html_content += (
                                    f"<td rowspan='{rowspan}' colspan='{colspan}'>"
                                )
                                html_content += (
                                    f"{cell_value if cell_value is not None else ''}"
                                )
                                html_content += "</td>"
                        else:
                            # 如果不是合并单元格的一部分，正常输出
                            if current_pos not in merged_cell_map:
                                html_content += f"<td>{cell_value if cell_value is not None else ''}</td>"

                    html_content += "</tr>"

                html_content += "</table><br>"

            except Exception as sheet_error:
                print(f"Error processing sheet {sheet_name}: {sheet_error}")
                html_content += f"<p>Error processing sheet {sheet_name}</p>"

        html_content += "</body></html>"
        return html_content

    except FileNotFoundError:
        print(f"Error: Excel file '{excel_file}' not found.")
        return "<html><body><p>Excel file not found.</p></body></html>"
    except PermissionError:
        print(f"Error: Permission denied when accessing '{excel_file}'.")
        return "<html><body><p>Permission denied accessing the file.</p></body></html>"
    except Exception as e:
        print(f"Unexpected error in xl_to_html: {e}")
        return f"<html><body><p>An unexpected error occurred: {e}</p></body></html>"


def html_table_to_excel_complex(
    html_content=None, filename="output.xlsx", table_index=0
):
    try:
        if html_content is None:
            raise ValueError("必须提供html_content参数")

        # 使用BeautifulSoup解析HTML
        try:
            soup = BeautifulSoup(html_content, "html.parser")
        except Exception as parse_error:
            print(f"HTML解析错误: {parse_error}")
            return False

        # 找到所有表格
        tables = soup.find_all("table")

        if not tables:
            print("未找到任何表格")
            return False

        if table_index >= len(tables):
            print(f"表格索引超出范围，共找到{len(tables)}个表格")
            return False

        # 选择指定索引的表格
        target_table = tables[table_index]

        # 创建工作簿和工作表
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
        except Exception as workbook_error:
            print(f"创建工作簿时出错: {workbook_error}")
            return False

        # 创建一个二维矩阵来跟踪哪些单元格已被填充
        rows = target_table.find_all("tr")
        max_cols = max([len(row.find_all(["th", "td"])) for row in rows])
        filled_cells = [
            [False for _ in range(max_cols + 10)] for _ in range(len(rows) + 10)
        ]

        current_row = 1

        try:
            for i, row in enumerate(rows):
                cells = row.find_all(["th", "td"])
                current_col = 1

                for cell in cells:
                    # 找到下一个未填充的单元格位置
                    while (
                        current_col <= max_cols + 5
                        and filled_cells[current_row - 1][current_col - 1]
                    ):
                        current_col += 1

                    # 处理rowspan和colspan
                    try:
                        rowspan = int(cell.get("rowspan", 1))
                        colspan = int(cell.get("colspan", 1))
                    except (ValueError, TypeError):
                        rowspan = 1
                        colspan = 1

                    # 获取单元格内容
                    value = cell.get_text(strip=True)

                    try:
                        # 写入单元格
                        ws.cell(row=current_row, column=current_col, value=value)

                        # 标记已填充的单元格
                        for r in range(rowspan):
                            for c in range(colspan):
                                if current_row - 1 + r < len(
                                    filled_cells
                                ) and current_col - 1 + c < len(filled_cells[0]):
                                    filled_cells[current_row - 1 + r][
                                        current_col - 1 + c
                                    ] = True

                        # 处理合并单元格
                        if rowspan > 1 or colspan > 1:
                            try:
                                ws.merge_cells(
                                    start_row=current_row,
                                    start_column=current_col,
                                    end_row=current_row + rowspan - 1,
                                    end_column=current_col + colspan - 1,
                                )
                            except Exception as merge_error:
                                print(f"合并单元格错误: {merge_error}")

                        # 移动到下一列位置
                        current_col += colspan

                    except IllegalCharacterError:
                        print(f"警告: 单元格({current_row},{current_col})包含不合法的字符，已替换")
                        ws.cell(row=current_row, column=current_col, value="[不支持的字符]")
                    except Exception as cell_error:
                        print(f"处理单元格({current_row},{current_col})时出错: {cell_error}")

                # 移动到下一行
                current_row += 1

        except Exception as row_process_error:
            print(f"处理行时出错: {row_process_error}")
            return False

        # 保存文件
        try:
            # 确保目录存在
            os.makedirs(os.path.dirname(os.path.abspath(filename)), exist_ok=True)

            wb.save(filename)
            print(f"Excel文件已成功保存到: {os.path.abspath(filename)}")
            return True

        except PermissionError:
            print(f"错误：没有权限保存文件 {filename}")
            return False
        except Exception as save_error:
            print(f"保存Excel文件时出错: {save_error}")
            return False

    except Exception as global_error:
        print(f"发生未预料的错误: {global_error}")
        return False


def convert_html_txt_to_dict(txt_file_path, dict_html):
    import re

    try:
        with open(txt_file_path, "r", encoding="utf-8") as file:
            for line in file:
                line = line.strip()
                if not line:
                    continue
                parts = re.split(r"\s+", line, 1)
                if len(parts) == 2:
                    filename = parts[0]
                    html_content = parts[1]
                    dict_html[filename] = html_content
                else:
                    print(f"警告：无法解析行: {line}")
    except FileNotFoundError:
        print(f"错误：找不到文件 {txt_file_path}")
    except Exception as e:
        print(f"错误：{str(e)}")


def save_dict_to_html_txt(dict_html, txt_file_path):
    try:
        # 打开文件用于写入，使用UTF-8编码
        with open(txt_file_path, "w", encoding="utf-8") as file:
            # 遍历字典中的所有键值对
            for filename, html_content in dict_html.items():
                # 写入格式：文件名 + 制表符 + HTML内容 + 换行符
                file.write(f"{filename}\t{html_content}\n")

        print(f"成功写入到预标注文件: {txt_file_path}")
        return True

    except Exception as e:
        print(f"写入预标注文件时发生错误: {str(e)}")
        return False
