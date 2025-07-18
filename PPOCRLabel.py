# Copyright (c) <2015-Present> Tzutalin
# Copyright (C) 2013  MIT, Computer Science and Artificial Intelligence Laboratory. Bryan Russell, Antonio Torralba,
# William T. Freeman. Permission is hereby granted, free of charge, to any person obtaining a copy of this software and
# associated documentation files (the "Software"), to deal in the Software without restriction, including without
# limitation the rights to use, copy, modify, merge, publish, distribute, sublicense, and/or sell copies of the
# Software, and to permit persons to whom the Software is furnished to do so, subject to the following conditions:
# The above copyright notice and this permission notice shall be included in all copies or substantial portions of
# the Software. THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR IMPLIED, INCLUDING BUT
# NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY, FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT
# SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF
# CONTRACT, TORT OR OTHERWISE, ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN
# THE SOFTWARE.
# !/usr/bin/env python
# -*- coding: utf-8 -*-
# pyrcc5 -o libs/resources.py resources.qrc
import argparse
import ast
import codecs
import json
import os
import platform
import subprocess
import sys
from functools import partial

import openpyxl
import cv2
import numpy as np

from PyQt5.QtCore import (
    QSize,
    Qt,
    QPoint,
    QByteArray,
    QTimer,
    QFileInfo,
    QPointF,
    QProcess,
)
from PyQt5.QtGui import (
    QImage,
    QCursor,
    QPixmap,
    QImageReader,
    QColor,
    QIcon,
    QFontDatabase,
)
from PyQt5.QtWidgets import (
    QMainWindow,
    QListWidget,
    QVBoxLayout,
    QSpinBox,
    QToolButton,
    QHBoxLayout,
    QDockWidget,
    QWidget,
    QSlider,
    QGraphicsOpacityEffect,
    QMessageBox,
    QListView,
    QScrollArea,
    QWidgetAction,
    QApplication,
    QLabel,
    QGridLayout,
    QFileDialog,
    QListWidgetItem,
    QComboBox,
    QDialog,
    QAbstractItemView,
    QMenu,
    QAction,
    QPushButton,
)

__dir__ = os.path.dirname(__file__)

from pandas.io.sql import has_table

sys.path.append(os.path.join(__dir__, ""))

import paddle
from paddleocr import PaddleOCR, PPStructureV3, TextRecognition, TextDetection
import libs.resources
from libs.constants import (
    SETTING_ADVANCE_MODE,
    SETTING_DRAW_SQUARE,
    SETTING_FILENAME,
    SETTING_FILL_COLOR,
    SETTING_LAST_OPEN_DIR,
    SETTING_LINE_COLOR,
    SETTING_PAINT_INDEX,
    SETTING_PAINT_LABEL,
    SETTING_RECENT_FILES,
    SETTING_SAVE_DIR,
    SETTING_WIN_POSE,
    SETTING_WIN_SIZE,
    SETTING_WIN_STATE,
)
from libs.utils import (
    addActions,
    boxPad,
    convert_token,
    expand_list,
    fmtShortcut,
    get_rotate_crop_image,
    have_qstring,
    keysInfo,
    natural_sort,
    newAction,
    newIcon,
    rebuild_html_from_ppstructure_label,
    stepsInfo,
    polygon_bounding_box_center_and_area,
    map_value,
    struct,
)
from libs.labelColor import label_colormap
from libs.settings import Settings
from libs.shape import Shape, DEFAULT_LINE_COLOR, DEFAULT_FILL_COLOR, DEFAULT_LOCK_COLOR
from libs.stringBundle import StringBundle
from libs.canvas import Canvas
from libs.zoomWidget import ZoomWidget
from libs.autoDialog import AutoDialog
from libs.labelDialog import LabelDialog
from libs.colorDialog import ColorDialog
from libs.hashableQListWidgetItem import HashableQListWidgetItem
from libs.editinlist import EditInList
from libs.unique_label_qlist_widget import UniqueLabelQListWidget
from libs.keyDialog import KeyDialog
from tablepyxl import tablepyxl

import logging
from datetime import datetime

logger = logging.getLogger("PPOCRLabel")


__appname__ = "PPOCRLabel"

LABEL_COLORMAP = label_colormap()


class MainWindow(QMainWindow):
    FIT_WINDOW, FIT_WIDTH, MANUAL_ZOOM = list(range(3))

    def __init__(
        self,
        lang="ch",
        gpu=False,
        img_list_natural_sort=True,
        bbox_auto_zoom_center=False,
        kie_mode=False,
        default_filename=None,
        default_predefined_class_file=None,
        default_save_dir=None,
        det_model_dir=None,
        rec_model_dir=None,
        cls_model_dir=None,
        label_font_path=None,
        selected_shape_color=(255, 255, 0),
    ):
        super(MainWindow, self).__init__()
        self.setWindowTitle(__appname__)
        self.setWindowState(Qt.WindowMaximized)  # set window max
        self.activateWindow()  # PPOCRLabel goes to the front when activate
        # gsj person use
        self.dict_html = {}
        self.dict_excel = {}
        self.dict_export = {}

        # Load setting in the main thread
        self.settings = Settings()
        self.settings.load()
        settings = self.settings
        self.lang = lang
        self.gpu = "gpu" if paddle.is_compiled_with_cuda() and gpu else "cpu"
        self.img_list_natural_sort = img_list_natural_sort
        self.bbox_auto_zoom_center = bbox_auto_zoom_center

        # Load string bundle for i18n
        if lang not in ["ch", "en"]:
            lang = "en"
        self.stringBundle = StringBundle.getBundle(
            localeStr="zh-CN" if lang == "ch" else "en"
        )  # 'en'

        def get_str(str_id):
            return self.stringBundle.getString(str_id)

        # KIE setting
        self.kie_mode = kie_mode
        self.key_previous_text = ""
        self.existed_key_cls_set = set()
        self.key_dialog_tip = get_str("keyDialogTip")

        self.defaultSaveDir = default_save_dir

        params = {
            "use_doc_orientation_classify": False,
            "use_doc_unwarping": False,
            "use_textline_orientation": True,
            "device": self.gpu,
            "lang": self.lang,
            "text_detection_model_name": "PP-OCRv5_server_det",
            "text_recognition_model_name": "PP-OCRv5_server_rec",
            "enable_mkldnn": False,
        }

        if det_model_dir is not None:
            params["text_detection_model_dir"] = det_model_dir
        if rec_model_dir is not None:
            params["text_recognition_model_dir"] = rec_model_dir
        if cls_model_dir is not None:
            params["text_line_orientation_model_dir"] = cls_model_dir

        self.ocr = PaddleOCR(**params)
        self.text_recognizer = TextRecognition(
            model_name="PP-OCRv5_server_rec",
            model_dir=rec_model_dir,
            device=self.gpu,
        )
        self.text_detector = TextDetection(
            model_name="PP-OCRv5_server_det",
            model_dir=det_model_dir,
            device=self.gpu,
        )
        self.table_ocr = PPStructureV3(
            use_doc_orientation_classify=False,
            use_doc_unwarping=False,
            use_seal_recognition=False,
            use_table_recognition=True,
            use_formula_recognition=False,
            use_chart_recognition=False,
            use_region_detection=False,
            device=self.gpu,
        )

        if os.path.exists("./data/paddle.png"):
            self.ocr.predict("./data/paddle.png")
            self.table_ocr.predict("./data/paddle.png")

        # For loading all image under a directory
        self.mImgList = []
        self.mImgList5 = []
        self.dirname = None
        self.labelHist = []
        self.lastOpenDir = None
        self.result_dic = []
        self.result_dic_locked = []
        self.changeFileFolder = False
        self.haveAutoReced = False
        self.labelFile = None
        self.currIndex = 0

        # Whether we need to save or not.
        self.dirty = False

        self._noSelectionSlot = False
        self._beginner = True
        self.screencastViewer = self.getAvailableScreencastViewer()
        self.screencast = "https://github.com/PFCCLab/PPOCRLabel"

        # Load predefined classes to the list
        self.loadPredefinedClasses(default_predefined_class_file)

        # Main widgets and related state.
        self.labelDialog = LabelDialog(parent=self, listItem=self.labelHist)
        self.autoDialog = AutoDialog(parent=self)

        self.itemsToShapes = {}
        self.shapesToItems = {}
        self.itemsToShapesbox = {}
        self.shapesToItemsbox = {}
        self.prevLabelText = get_str("tempLabel")
        self.noLabelText = get_str("nullLabel")
        self.model = "paddle"
        self.PPreader = None
        self.autoSaveNum = 5

        #  ================== File List  ==================

        filelistLayout = QVBoxLayout()
        filelistLayout.setContentsMargins(0, 0, 0, 0)

        self.fileListWidget = QListWidget()
        self.fileListWidget.itemClicked.connect(self.fileitemDoubleClicked)
        self.fileListWidget.setIconSize(QSize(25, 25))
        filelistLayout.addWidget(self.fileListWidget)

        fileListContainer = QWidget()
        fileListContainer.setLayout(filelistLayout)
        self.fileListName = get_str("fileList")
        self.fileDock = QDockWidget(self.fileListName, self)
        self.fileDock.setObjectName(get_str("files"))
        self.fileDock.setWidget(fileListContainer)
        self.addDockWidget(Qt.LeftDockWidgetArea, self.fileDock)

        #  ================== Key List  ==================
        if self.kie_mode:
            self.keyList = UniqueLabelQListWidget()

            # set key list height
            key_list_height = int(QApplication.desktop().height() // 4)
            if key_list_height < 50:
                key_list_height = 50
            self.keyList.setMaximumHeight(key_list_height)

            self.keyListDockName = get_str("keyListTitle")
            self.keyListDock = QDockWidget(self.keyListDockName, self)
            self.keyListDock.setWidget(self.keyList)
            self.keyListDock.setFeatures(QDockWidget.NoDockWidgetFeatures)
            filelistLayout.addWidget(self.keyListDock)

        self.auto_recognition_num = 1

        self.AutoRecognitionNum = QSpinBox()
        self.AutoRecognitionNum.valueChanged.connect(self.autoRecognitionNum)
        self.AutoRecognitionNum.setFixedWidth(80)

        self.AutoRecognition = QToolButton()
        self.AutoRecognition.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.AutoRecognition.setIcon(newIcon("Auto"))
        autoRecLayout = QHBoxLayout()
        autoRecLayout.setContentsMargins(0, 0, 0, 0)
        autoRecLayout.addWidget(self.AutoRecognitionNum)
        autoRecLayout.addWidget(self.AutoRecognition)
        autoRecContainer = QWidget()
        autoRecContainer.setLayout(autoRecLayout)
        filelistLayout.addWidget(autoRecContainer)

        #  ================== Right Area  ==================
        listLayout = QVBoxLayout()
        listLayout.setContentsMargins(0, 0, 0, 0)

        # Buttons
        self.editButton = QToolButton()
        self.reRecogButton = QToolButton()
        self.reRecogButton.setIcon(newIcon("reRec", 30))
        self.reRecogButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

        self.tableRecButton = QToolButton()
        self.tableRecButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

        self.newButton = QToolButton()
        self.newButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.createpolyButton = QToolButton()
        self.createpolyButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

        self.SaveButton = QToolButton()
        self.SaveButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.DelButton = QToolButton()
        self.DelButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.ResortButton = QToolButton()
        self.ResortButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.ImportButton = QToolButton()
        self.ImportButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.ExportButton = QToolButton()
        self.ExportButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)

        leftTopToolBox = QGridLayout()
        leftTopToolBox.addWidget(self.newButton, 0, 0, 1, 1)
        leftTopToolBox.addWidget(self.createpolyButton, 0, 1, 1, 1)
        leftTopToolBox.addWidget(self.reRecogButton, 1, 0, 1, 1)
        leftTopToolBox.addWidget(self.tableRecButton, 1, 1, 1, 1)
        leftTopToolBox.addWidget(self.tableRecButton, 1, 1, 1, 1)
        leftTopToolBox.addWidget(self.ImportButton, 0, 2, 1, 1)
        leftTopToolBox.addWidget(self.ExportButton, 1, 2, 1, 1)

        leftTopToolBoxContainer = QWidget()
        leftTopToolBoxContainer.setLayout(leftTopToolBox)
        listLayout.addWidget(leftTopToolBoxContainer)

        #  ================== Label List  ==================
        labelIndexListlBox = QHBoxLayout()

        # Create and add a widget for showing current label item index
        self.indexList = QListWidget()
        self.indexList.setMaximumSize(30, 16777215)  # limit max width
        self.indexList.setEditTriggers(QAbstractItemView.NoEditTriggers)  # no editable
        self.indexList.itemSelectionChanged.connect(self.indexSelectionChanged)
        self.indexList.setVerticalScrollBarPolicy(
            Qt.ScrollBarAlwaysOff
        )  # no scroll Bar
        self.indexListDock = QDockWidget("No.", self)
        self.indexListDock.setWidget(self.indexList)
        self.indexListDock.setFeatures(QDockWidget.NoDockWidgetFeatures)
        labelIndexListlBox.addWidget(self.indexListDock, 1)
        # no margin between two boxes
        labelIndexListlBox.setSpacing(0)

        # Create and add a widget for showing current label items
        self.labelList = EditInList()
        labelListContainer = QWidget()
        labelListContainer.setLayout(listLayout)
        self.labelList.itemSelectionChanged.connect(self.labelSelectionChanged)
        self.labelList.clicked.connect(self.labelList.item_clicked)

        # Connect to itemChanged to detect checkbox changes.
        self.labelList.itemChanged.connect(self.labelItemChanged)
        self.labelListDockName = get_str("recognitionResult")
        self.labelListDock = QDockWidget(self.labelListDockName, self)
        self.labelListDock.setWidget(self.labelList)
        self.labelListDock.setFeatures(QDockWidget.NoDockWidgetFeatures)
        labelIndexListlBox.addWidget(
            self.labelListDock, 10
        )  # label list is wider than index list

        # enable labelList drag_drop to adjust bbox order
        # Set selection mode to single selection
        self.labelList.setSelectionMode(QAbstractItemView.SingleSelection)
        # Enable drag functionality
        self.labelList.setDragEnabled(True)
        # Set to accept drops
        self.labelList.viewport().setAcceptDrops(True)
        # Show drop indicator position
        self.labelList.setDropIndicatorShown(True)
        # Set drag-drop mode to move items, if not set, default is copy items
        self.labelList.setDragDropMode(QAbstractItemView.InternalMove)
        # Trigger drop event
        self.labelList.model().rowsMoved.connect(self.drag_drop_happened)

        labelIndexListContainer = QWidget()
        labelIndexListContainer.setLayout(labelIndexListlBox)
        listLayout.addWidget(labelIndexListContainer)

        # Synchronize scrolling between labelList and indexList
        self.labelListBar = self.labelList.verticalScrollBar()
        self.indexListBar = self.indexList.verticalScrollBar()

        self.labelListBar.valueChanged.connect(self.move_scrollbar)
        self.indexListBar.valueChanged.connect(self.move_scrollbar)

        #  ================== Detection Box  ==================
        self.BoxList = QListWidget()

        # self.BoxList.itemActivated.connect(self.boxSelectionChanged)
        self.BoxList.itemSelectionChanged.connect(self.boxSelectionChanged)
        self.BoxList.itemDoubleClicked.connect(self.editBox)
        # Connect to itemChanged to detect checkbox changes.
        self.BoxList.itemChanged.connect(self.boxItemChanged)
        self.BoxListDockName = get_str("detectionBoxposition")
        self.BoxListDock = QDockWidget(self.BoxListDockName, self)
        self.BoxListDock.setWidget(self.BoxList)
        self.BoxListDock.setFeatures(QDockWidget.NoDockWidgetFeatures)
        listLayout.addWidget(self.BoxListDock)

        #  ================== Lower Right Area  ==================
        leftbtmtoolbox = QHBoxLayout()
        leftbtmtoolbox.addWidget(self.SaveButton)
        leftbtmtoolbox.addWidget(self.DelButton)
        leftbtmtoolbox.addWidget(self.ResortButton)
        leftbtmtoolboxcontainer = QWidget()
        leftbtmtoolboxcontainer.setLayout(leftbtmtoolbox)
        listLayout.addWidget(leftbtmtoolboxcontainer)

        self.dock = QDockWidget(get_str("boxLabelText"), self)
        self.dock.setObjectName(get_str("labels"))
        self.dock.setWidget(labelListContainer)

        #  ================== Zoom Bar  ==================
        self.imageSlider = QSlider(Qt.Horizontal)
        self.imageSlider.valueChanged.connect(self.CanvasSizeChange)
        self.imageSlider.setMinimum(-9)
        self.imageSlider.setMaximum(510)
        self.imageSlider.setSingleStep(1)
        self.imageSlider.setTickPosition(QSlider.TicksBelow)
        self.imageSlider.setTickInterval(1)

        op = QGraphicsOpacityEffect()
        op.setOpacity(0.2)
        self.imageSlider.setGraphicsEffect(op)

        self.imageSlider.setStyleSheet("background-color:transparent")
        self.imageSliderDock = QDockWidget(get_str("ImageResize"), self)
        self.imageSliderDock.setObjectName(get_str("IR"))
        self.imageSliderDock.setWidget(self.imageSlider)
        self.imageSliderDock.setFeatures(QDockWidget.DockWidgetFloatable)
        self.imageSliderDock.setAttribute(Qt.WA_TranslucentBackground)
        self.addDockWidget(Qt.RightDockWidgetArea, self.imageSliderDock)

        self.zoomWidget = ZoomWidget()
        self.colorDialog = ColorDialog(parent=self)
        self.zoomWidgetValue = self.zoomWidget.value()

        self.msgBox = QMessageBox()

        #  ================== Thumbnail ==================
        hlayout = QHBoxLayout()
        m = (0, 0, 0, 0)
        hlayout.setSpacing(0)
        hlayout.setContentsMargins(*m)
        self.preButton = QToolButton()
        self.preButton.setIcon(newIcon("prev", 40))
        self.preButton.setIconSize(QSize(40, 100))
        self.preButton.clicked.connect(self.openPrevImg)
        self.preButton.setStyleSheet("border: none;")
        self.preButton.setShortcut("a")
        self.iconlist = QListWidget()
        self.iconlist.setViewMode(QListView.IconMode)
        self.iconlist.setFlow(QListView.TopToBottom)
        self.iconlist.setSpacing(10)
        self.iconlist.setIconSize(QSize(50, 50))
        self.iconlist.setMovement(QListView.Static)
        self.iconlist.setResizeMode(QListView.Adjust)
        self.iconlist.itemClicked.connect(self.iconitemDoubleClicked)
        self.iconlist.setStyleSheet(
            "QListWidget{ background-color:transparent; border: none;}"
        )
        self.iconlist.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.nextButton = QToolButton()
        self.nextButton.setIcon(newIcon("next", 40))
        self.nextButton.setIconSize(QSize(40, 100))
        self.nextButton.setStyleSheet("border: none;")
        self.nextButton.clicked.connect(self.openNextImg)
        self.nextButton.setShortcut("d")

        hlayout.addWidget(self.preButton)
        hlayout.addWidget(self.iconlist)
        hlayout.addWidget(self.nextButton)

        iconListContainer = QWidget()
        iconListContainer.setLayout(hlayout)
        iconListContainer.setFixedHeight(100)

        #  ================== Canvas ==================
        self.canvas = Canvas(parent=self)
        self.canvas.zoomRequest.connect(self.zoomRequest)
        self.canvas.setDrawingShapeToSquare(settings.get(SETTING_DRAW_SQUARE, False))

        scroll = QScrollArea()
        scroll.setWidget(self.canvas)
        scroll.setWidgetResizable(True)
        self.scrollBars = {
            Qt.Vertical: scroll.verticalScrollBar(),
            Qt.Horizontal: scroll.horizontalScrollBar(),
        }
        self.scrollArea = scroll
        self.canvas.scrollRequest.connect(self.scrollRequest)

        self.canvas.newShape.connect(partial(self.newShape, False))
        self.canvas.shapeMoved.connect(self.updateBoxlist)  # self.setDirty
        self.canvas.selectionChanged.connect(self.shapeSelectionChanged)
        self.canvas.drawingPolygon.connect(self.toggleDrawingSensitive)

        centerLayout = QVBoxLayout()
        centerLayout.setContentsMargins(0, 0, 0, 0)
        centerLayout.addWidget(scroll)
        centerLayout.addWidget(iconListContainer, 0, Qt.AlignCenter)
        centerContainer = QWidget()
        centerContainer.setLayout(centerLayout)

        self.setCentralWidget(centerContainer)
        self.addDockWidget(Qt.RightDockWidgetArea, self.dock)

        self.dock.setFeatures(
            QDockWidget.DockWidgetClosable | QDockWidget.DockWidgetFloatable
        )
        self.fileDock.setFeatures(QDockWidget.NoDockWidgetFeatures)

        #  ================== Actions ==================
        action = partial(newAction, self)
        quit = action(get_str("quit"), self.close, "Ctrl+Q", "quit", get_str("quitApp"))

        opendir = action(
            get_str("openDir"), self.openDirDialog, "Ctrl+u", "open", get_str("openDir")
        )

        open_dataset_dir = action(
            get_str("openDatasetDir"),
            self.openDatasetDirDialog,
            "Ctrl+p",
            "open",
            get_str("openDatasetDir"),
            enabled=False,
        )

        save = action(
            get_str("save"),
            self.saveFile,
            ["Ctrl+V", "end"],
            "verify",
            get_str("saveDetail"),
            enabled=False,
        )

        alcm = action(
            get_str("choosemodel"),
            self.autolcm,
            "Ctrl+M",
            "next",
            get_str("tipchoosemodel"),
        )

        deleteImg = action(
            get_str("deleteImg"),
            self.deleteImg,
            "Ctrl+Shift+D",
            "close",
            get_str("deleteImgDetail"),
            enabled=True,
        )

        resetAll = action(
            get_str("resetAll"),
            self.resetAll,
            None,
            "resetall",
            get_str("resetAllDetail"),
        )

        color1 = action(
            get_str("boxLineColor"),
            self.chooseColor,
            "Ctrl+L",
            "color_line",
            get_str("boxLineColorDetail"),
        )

        createMode = action(
            get_str("crtBox"),
            self.setCreateMode,
            "w",
            "new",
            get_str("crtBoxDetail"),
            enabled=False,
        )
        editMode = action(
            "&Edit\nRectBox",
            self.setEditMode,
            "Ctrl+J",
            "edit",
            "Move and edit Boxs",
            enabled=False,
        )

        create = action(
            get_str("crtBox"),
            self.createShape,
            "w",
            "objects",
            get_str("crtBoxDetail"),
            enabled=False,
        )

        delete = action(
            get_str("delBox"),
            self.deleteSelectedShape,
            ["backspace", "delete"],
            "delete",
            get_str("delBoxDetail"),
            enabled=False,
        )

        copy = action(
            get_str("dupBox"),
            self.copySelectedShape,
            "Ctrl+C",
            "copy",
            get_str("dupBoxDetail"),
            enabled=False,
        )

        hideAll = action(
            get_str("hideBox"),
            partial(self.togglePolygons, False),
            "Ctrl+H",
            "hide",
            get_str("hideAllBoxDetail"),
            enabled=False,
        )
        showAll = action(
            get_str("showBox"),
            partial(self.togglePolygons, True),
            "Ctrl+A",
            "hide",
            get_str("showAllBoxDetail"),
            enabled=False,
        )

        help = action(
            get_str("tutorial"),
            self.showTutorialDialog,
            None,
            "help",
            get_str("tutorialDetail"),
        )
        showInfo = action(
            get_str("info"), self.showInfoDialog, None, "help", get_str("info")
        )
        showSteps = action(
            get_str("steps"), self.showStepsDialog, None, "help", get_str("steps")
        )
        showKeys = action(
            get_str("keys"), self.showKeysDialog, None, "help", get_str("keys")
        )

        zoom = QWidgetAction(self)
        zoom.setDefaultWidget(self.zoomWidget)
        self.zoomWidget.setWhatsThis(
            "Zoom in or out of the image. Also accessible with"
            " %s and %s from the canvas."
            % (fmtShortcut("Ctrl+[-+]"), fmtShortcut("Ctrl+Wheel"))
        )
        self.zoomWidget.setEnabled(False)

        zoomIn = action(
            get_str("zoomin"),
            partial(self.addZoom, 10),
            "Ctrl++",
            "zoom-in",
            get_str("zoominDetail"),
            enabled=False,
        )
        zoomOut = action(
            get_str("zoomout"),
            partial(self.addZoom, -10),
            "Ctrl+-",
            "zoom-out",
            get_str("zoomoutDetail"),
            enabled=False,
        )
        zoomOrg = action(
            get_str("originalsize"),
            partial(self.setZoom, 100),
            "Ctrl+=",
            "zoom",
            get_str("originalsizeDetail"),
            enabled=False,
        )
        fitWindow = action(
            get_str("fitWin"),
            self.setFitWindow,
            "Ctrl+F",
            "fit-window",
            get_str("fitWinDetail"),
            checkable=True,
            enabled=False,
        )
        fitWidth = action(
            get_str("fitWidth"),
            self.setFitWidth,
            "Ctrl+Shift+F",
            "fit-width",
            get_str("fitWidthDetail"),
            checkable=True,
            enabled=False,
        )
        # Group zoom controls into a list for easier toggling.
        zoomActions = (self.zoomWidget, zoomIn, zoomOut, zoomOrg, fitWindow, fitWidth)
        self.zoomMode = self.MANUAL_ZOOM
        self.scalers = {
            self.FIT_WINDOW: self.scaleFitWindow,
            self.FIT_WIDTH: self.scaleFitWidth,
            # Set to one to scale to 100% when loading files.
            self.MANUAL_ZOOM: lambda: 1,
        }

        #  ================== New Actions ==================

        edit = action(
            get_str("editLabel"),
            self.editLabel,
            "Ctrl+E",
            "edit",
            get_str("editLabelDetail"),
            enabled=False,
        )

        AutoRec = action(
            get_str("autoRecognition"),
            self.autoRecognition,
            "",
            "Auto",
            get_str("autoRecognition"),
            enabled=False,
        )

        reRec = action(
            get_str("reRecognition"),
            self.reRecognition,
            "Ctrl+Shift+R",
            "reRec",
            get_str("reRecognition"),
            enabled=False,
        )

        singleRere = action(
            get_str("singleRe"),
            self.singleRerecognition,
            "Ctrl+R",
            "reRec",
            get_str("singleRe"),
            enabled=False,
        )

        createpoly = action(
            get_str("creatPolygon"),
            self.createPolygon,
            ["q", "home"],
            "new",
            get_str("creatPolygon"),
            enabled=False,
        )

        tableRec = action(
            get_str("TableRecognition"),
            self.TableRecognition,
            "",
            "Auto",
            get_str("TableRecognition"),
            enabled=False,
        )

        cellreRec = action(
            get_str("cellreRecognition"),
            self.cellreRecognition,
            "",
            "reRec",
            get_str("cellreRecognition"),
            enabled=False,
        )

        saveRec = action(
            get_str("saveRec"),
            self.saveRecResult,
            "",
            "save",
            get_str("saveRec"),
            enabled=False,
        )

        saveLabel = action(
            get_str("saveLabel"),
            self.saveLabelFile,  #
            "Ctrl+S",
            "save",
            get_str("saveLabel"),
            enabled=False,
        )

        exportJSON = action(
            get_str("exportJSON"),
            self.exportJSON,
            "",
            "save",
            get_str("exportJSON"),
            enabled=False,
        )

        undoLastPoint = action(
            get_str("undoLastPoint"),
            self.canvas.undoLastPoint,
            "Ctrl+Z",
            "undo",
            get_str("undoLastPoint"),
            enabled=False,
        )

        rotateLeft = action(
            get_str("rotateLeft"),
            partial(self.rotateImgAction, 1),
            "Ctrl+Alt+L",
            "rotateLeft",
            get_str("rotateLeft"),
            enabled=False,
        )

        rotateRight = action(
            get_str("rotateRight"),
            partial(self.rotateImgAction, -1),
            "Ctrl+Alt+R",
            "rotateRight",
            get_str("rotateRight"),
            enabled=False,
        )

        undo = action(
            get_str("undo"),
            self.undoShapeEdit,
            "Ctrl+Z",
            "undo",
            get_str("undo"),
            enabled=False,
        )

        change_cls = action(
            get_str("keyChange"),
            self.change_box_key,
            "Ctrl+X",
            "edit",
            get_str("keyChange"),
            enabled=False,
        )

        lock = action(
            get_str("lockBox"),
            self.lockSelectedShape,
            None,
            "lock",
            get_str("lockBoxDetail"),
            enabled=False,
        )
        expand = action(
            get_str("expandBox"),
            self.expandSelectedShape,
            "Ctrl+K",
            "expand",
            get_str("expandBoxDetail"),
            enabled=False,
        )
        resort = action(
            get_str("resortposition"),
            self.resortBoxPosition,
            "Ctrl+B",
            "resort",
            get_str("resortpositiondetail"),
            enabled=True,
        )
        importhtml = action(
            get_str("importhtml"),
            self.importhtml,
            "Ctrl+I",
            "importhtml",
            get_str("importhtmldetail"),
            enabled=True,
        )
        exporthtml = action(
            get_str("exporthtml"),
            self.exporthtml,
            "Ctrl+E",
            "exporthtml",
            get_str("exporthtmldetail"),
            enabled=True,
        )

        self.editButton.setDefaultAction(edit)
        self.newButton.setDefaultAction(create)
        self.createpolyButton.setDefaultAction(createpoly)
        self.DelButton.setDefaultAction(deleteImg)
        self.SaveButton.setDefaultAction(save)
        self.AutoRecognition.setDefaultAction(AutoRec)
        self.reRecogButton.setDefaultAction(reRec)
        self.tableRecButton.setDefaultAction(tableRec)
        self.ResortButton.setDefaultAction(resort)
        self.ImportButton.setDefaultAction(importhtml)
        self.ExportButton.setDefaultAction(exporthtml)
        # self.preButton.setDefaultAction(openPrevImg)
        # self.nextButton.setDefaultAction(openNextImg)

        #  ================== Zoom layout ==================
        zoomLayout = QHBoxLayout()
        zoomLayout.addStretch()
        self.zoominButton = QToolButton()
        self.zoominButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.zoominButton.setDefaultAction(zoomIn)
        self.zoomoutButton = QToolButton()
        self.zoomoutButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.zoomoutButton.setDefaultAction(zoomOut)
        self.zoomorgButton = QToolButton()
        self.zoomorgButton.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.zoomorgButton.setDefaultAction(zoomOrg)
        zoomLayout.addWidget(self.zoominButton)
        zoomLayout.addWidget(self.zoomorgButton)
        zoomLayout.addWidget(self.zoomoutButton)

        zoomContainer = QWidget()
        zoomContainer.setLayout(zoomLayout)
        zoomContainer.setGeometry(0, 0, 30, 150)

        shapeLineColor = action(
            get_str("shapeLineColor"),
            self.chshapeLineColor,
            icon="color_line",
            tip=get_str("shapeLineColorDetail"),
            enabled=False,
        )
        shapeFillColor = action(
            get_str("shapeFillColor"),
            self.chshapeFillColor,
            icon="color",
            tip=get_str("shapeFillColorDetail"),
            enabled=False,
        )

        # Label list context menu.
        labelMenu = QMenu()
        addActions(labelMenu, (edit, delete))

        self.labelList.setContextMenuPolicy(Qt.CustomContextMenu)
        self.labelList.customContextMenuRequested.connect(self.popLabelListMenu)

        # Draw squares/rectangles
        self.drawSquaresOption = QAction(get_str("drawSquares"), self)
        self.drawSquaresOption.setCheckable(True)
        self.drawSquaresOption.setChecked(settings.get(SETTING_DRAW_SQUARE, False))
        self.drawSquaresOption.triggered.connect(self.toogleDrawSquare)

        # Store actions for further handling.
        self.actions = struct(
            save=save,
            resetAll=resetAll,
            deleteImg=deleteImg,
            lineColor=color1,
            create=create,
            createpoly=createpoly,
            tableRec=tableRec,
            delete=delete,
            edit=edit,
            copy=copy,
            saveRec=saveRec,
            singleRere=singleRere,
            AutoRec=AutoRec,
            reRec=reRec,
            cellreRec=cellreRec,
            createMode=createMode,
            editMode=editMode,
            shapeLineColor=shapeLineColor,
            shapeFillColor=shapeFillColor,
            zoom=zoom,
            zoomIn=zoomIn,
            zoomOut=zoomOut,
            zoomOrg=zoomOrg,
            fitWindow=fitWindow,
            fitWidth=fitWidth,
            zoomActions=zoomActions,
            saveLabel=saveLabel,
            change_cls=change_cls,
            undo=undo,
            undoLastPoint=undoLastPoint,
            open_dataset_dir=open_dataset_dir,
            rotateLeft=rotateLeft,
            rotateRight=rotateRight,
            lock=lock,
            exportJSON=exportJSON,
            expand=expand,
            resort=resort,
            fileMenuActions=(
                opendir,
                open_dataset_dir,
                saveLabel,
                exportJSON,
                resetAll,
                quit,
            ),
            beginner=(),
            advanced=(),
            editMenu=(
                createpoly,
                edit,
                copy,
                delete,
                singleRere,
                cellreRec,
                resort,
                None,
                undo,
                undoLastPoint,
                None,
                rotateLeft,
                rotateRight,
                None,
                color1,
                self.drawSquaresOption,
                lock,
                expand,
                None,
                change_cls,
            ),
            beginnerContext=(
                create,
                createpoly,
                edit,
                copy,
                delete,
                singleRere,
                cellreRec,
                rotateLeft,
                rotateRight,
                lock,
                expand,
                change_cls,
            ),
            advancedContext=(
                createMode,
                editMode,
                edit,
                copy,
                delete,
                shapeLineColor,
                shapeFillColor,
            ),
            onLoadActive=(create, createpoly, createMode, editMode),
            onShapesPresent=(hideAll, showAll),
        )

        # menus
        self.menus = struct(
            file=self.menu("&" + get_str("mfile")),
            edit=self.menu("&" + get_str("medit")),
            view=self.menu("&" + get_str("mview")),
            autolabel=self.menu("&PaddleOCR"),
            help=self.menu("&" + get_str("mhelp")),
            recentFiles=QMenu("Open &Recent"),
            labelList=labelMenu,
        )

        self.lastLabel = None
        # Add option to enable/disable labels being displayed at the top of bounding boxes
        self.displayLabelOption = QAction(get_str("displayLabel"), self)
        self.displayLabelOption.setShortcut("Ctrl+Shift+P")
        self.displayLabelOption.setCheckable(True)
        self.displayLabelOption.setChecked(settings.get(SETTING_PAINT_LABEL, False))
        self.displayLabelOption.triggered.connect(self.togglePaintLabelsOption)

        # Add option to enable/disable box index being displayed at the top of bounding boxes
        self.displayIndexOption = QAction(get_str("displayIndex"), self)
        self.displayIndexOption.setCheckable(True)
        self.displayIndexOption.setChecked(settings.get(SETTING_PAINT_INDEX, False))
        self.displayIndexOption.triggered.connect(self.togglePaintIndexOption)

        self.labelDialogOption = QAction(get_str("labelDialogOption"), self)
        self.labelDialogOption.setShortcut("Ctrl+Shift+L")
        self.labelDialogOption.setCheckable(True)
        self.labelDialogOption.setChecked(settings.get(SETTING_PAINT_LABEL, False))
        self.displayIndexOption.setChecked(settings.get(SETTING_PAINT_INDEX, False))
        self.labelDialogOption.triggered.connect(self.speedChoose)

        self.autoSaveOption = QAction(get_str("autoSaveMode"), self)
        self.autoSaveOption.setCheckable(True)
        self.autoSaveOption.setChecked(settings.get(SETTING_PAINT_LABEL, False))
        self.displayIndexOption.setChecked(settings.get(SETTING_PAINT_INDEX, False))
        self.autoSaveOption.triggered.connect(self.autoSaveFunc)

        self.autoImportOption = QAction(get_str("autoimporthtml"), self)
        self.autoImportOption.setCheckable(True)

        self.autoExportOption = QAction(get_str("autoexporthtml"), self)
        self.autoExportOption.setCheckable(True)

        self.autoCheck = QAction(get_str("autocheck"), self)
        self.autoCheck.setCheckable(True)

        self.autoReRecognitionOption = QAction(get_str("autoReRecognition"), self)
        self.autoReRecognitionOption.setCheckable(True)
        self.autoReRecognitionOption.setChecked(
            settings.get(SETTING_PAINT_LABEL, False)
        )
        self.displayIndexOption.setChecked(settings.get(SETTING_PAINT_INDEX, False))
        self.autoReRecognitionOption.triggered.connect(self.autoSaveFunc)

        self.autoSaveUnsavedChangesOption = QAction(
            get_str("autoSaveUnsavedChanges"), self
        )
        self.autoSaveUnsavedChangesOption.setCheckable(True)
        self.autoSaveUnsavedChangesOption.setChecked(
            settings.get(SETTING_PAINT_LABEL, False)
        )
        self.displayIndexOption.setChecked(settings.get(SETTING_PAINT_INDEX, False))
        self.autoSaveUnsavedChangesOption.triggered.connect(self.autoSaveFunc)

        addActions(
            self.menus.file,
            (
                opendir,
                open_dataset_dir,
                None,
                saveLabel,
                saveRec,
                exportJSON,
                self.autoSaveOption,
                self.autoReRecognitionOption,
                self.autoSaveUnsavedChangesOption,
                self.autoImportOption,
                self.autoExportOption,
                self.autoCheck,
                None,
                resetAll,
                deleteImg,
                quit,
            ),
        )

        addActions(self.menus.help, (showKeys, showSteps, showInfo))
        addActions(
            self.menus.view,
            (
                self.displayLabelOption,
                self.displayIndexOption,
                self.labelDialogOption,
                None,
                hideAll,
                showAll,
                None,
                zoomIn,
                zoomOut,
                zoomOrg,
                None,
                fitWindow,
                fitWidth,
            ),
        )

        addActions(self.menus.autolabel, (AutoRec, reRec, cellreRec, alcm, None, help))

        self.menus.file.aboutToShow.connect(self.updateFileMenu)

        # Custom context menu for the canvas widget:
        addActions(self.canvas.menus[0], self.actions.beginnerContext)

        self.statusBar().showMessage("%s started." % __appname__)
        self.statusBar().show()

        # Application state.
        self.image = QImage()
        self.filePath = default_filename
        self.lastOpenDir = None
        self.recentFiles = []
        self.maxRecent = 7
        self.lineColor = None
        self.fillColor = None
        self.zoom_level = 100
        self.fit_window = False
        # Add Chris
        self.difficult = False

        # Fix the compatible issue for qt4 and qt5. Convert the QStringList to python list
        if settings.get(SETTING_RECENT_FILES):
            if have_qstring():
                recentFileQStringList = settings.get(SETTING_RECENT_FILES)
                self.recentFiles = [i for i in recentFileQStringList]
            else:
                self.recentFiles = recentFileQStringList = settings.get(
                    SETTING_RECENT_FILES
                )

        size = settings.get(SETTING_WIN_SIZE, QSize(1200, 800))

        position = QPoint(0, 0)
        saved_position = settings.get(SETTING_WIN_POSE, position)
        # Fix the multiple monitors issue
        for i in range(QApplication.desktop().screenCount()):
            if QApplication.desktop().availableGeometry(i).contains(saved_position):
                position = saved_position
                break
        self.resize(size)
        self.move(position)
        saveDir = settings.get(SETTING_SAVE_DIR, None)
        logger.debug("Save directory: %s", saveDir)
        self.lastOpenDir = settings.get(SETTING_LAST_OPEN_DIR, None)

        self.restoreState(settings.get(SETTING_WIN_STATE, QByteArray()))
        Shape.line_color = self.lineColor = QColor(
            settings.get(SETTING_LINE_COLOR, DEFAULT_LINE_COLOR)
        )
        Shape.fill_color = self.fillColor = QColor(
            settings.get(SETTING_FILL_COLOR, DEFAULT_FILL_COLOR)
        )
        self.canvas.setDrawingColor(self.lineColor)
        # Add chris
        Shape.difficult = self.difficult

        # ADD:
        # Populate the File menu dynamically.
        self.updateFileMenu()

        # Since loading the file may take some time, make sure it runs in the background.
        if self.filePath and os.path.isdir(self.filePath):
            self.queueEvent(partial(self.importDirImages, self.filePath or ""))
        elif self.filePath:
            self.queueEvent(partial(self.loadFile, self.filePath or ""))

        self.keyDialog = None

        # Callbacks:
        self.zoomWidget.valueChanged.connect(self.paintCanvas)

        self.populateModeActions()

        # Display cursor coordinates at the right of status bar
        self.labelCoordinates = QLabel("")
        self.statusBar().addPermanentWidget(self.labelCoordinates)

        # Open Dir if deafult file
        if self.filePath and os.path.isdir(self.filePath):
            self.openDirDialog(dirpath=self.filePath, silent=True)

        # load label font
        self.label_font_family = None
        if label_font_path is not None:
            label_font_id = QFontDatabase.addApplicationFont(label_font_path)
            if label_font_id >= 0:
                self.label_font_family = QFontDatabase.applicationFontFamilies(
                    label_font_id
                )[0]

        # selected shape color
        self.selected_shape_color = selected_shape_color

    def menu(self, title, actions=None):
        menu = self.menuBar().addMenu(title)
        if actions:
            addActions(menu, actions)
        return menu

    def keyReleaseEvent(self, event):
        if event.key() == Qt.Key_Control:
            self.canvas.setDrawingShapeToSquare(False)

    def keyPressEvent(self, event):
        if event.key() == Qt.Key_Control:
            # Draw rectangle if Ctrl is pressed
            self.canvas.setDrawingShapeToSquare(True)

    def noShapes(self):
        return not self.itemsToShapes

    def populateModeActions(self):
        self.canvas.menus[0].clear()
        addActions(self.canvas.menus[0], self.actions.beginnerContext)
        self.menus.edit.clear()
        actions = (
            self.actions.create,
        )  # if self.beginner() else (self.actions.createMode, self.actions.editMode)
        addActions(self.menus.edit, actions + self.actions.editMenu)

    def setDirty(self):
        self.dirty = True
        self.actions.save.setEnabled(True)

    def setClean(self):
        self.dirty = False
        self.actions.save.setEnabled(False)
        self.actions.create.setEnabled(True)
        self.actions.createpoly.setEnabled(True)

    def toggleActions(self, value=True):
        """Enable/Disable widgets which depend on an opened image."""
        for z in self.actions.zoomActions:
            z.setEnabled(value)
        for action in self.actions.onLoadActive:
            action.setEnabled(value)

    def queueEvent(self, function):
        QTimer.singleShot(0, function)

    def status(self, message, delay=5000):
        self.statusBar().showMessage(message, delay)

    def resetState(self):
        self.itemsToShapes.clear()
        self.shapesToItems.clear()
        self.itemsToShapesbox.clear()  # ADD
        self.shapesToItemsbox.clear()
        self.labelList.clear()
        self.BoxList.clear()
        self.indexList.clear()
        self.filePath = None
        self.imageData = None
        self.labelFile = None
        self.canvas.resetState()
        self.labelCoordinates.clear()
        # self.comboBox.cb.clear()
        self.result_dic = []

    def currentItem(self):
        items = self.labelList.selectedItems()
        if items:
            return items[0]
        return None

    def currentBox(self):
        items = self.BoxList.selectedItems()
        if items:
            return items[0]
        return None

    def addRecentFile(self, filePath):
        if filePath in self.recentFiles:
            self.recentFiles.remove(filePath)
        elif len(self.recentFiles) >= self.maxRecent:
            self.recentFiles.pop()
        self.recentFiles.insert(0, filePath)

    def beginner(self):
        return self._beginner

    def advanced(self):
        return not self.beginner()

    def getAvailableScreencastViewer(self):
        osName = platform.system()

        if osName == "Windows":
            return ["C:\\Program Files\\Internet Explorer\\iexplore.exe"]
        elif osName == "Linux":
            return ["xdg-open"]
        elif osName == "Darwin":
            return ["open"]
        return [None]

    ## Callbacks ##
    def showTutorialDialog(self):
        subprocess.Popen(self.screencastViewer + [self.screencast])

    def showInfoDialog(self):
        from libs.__init__ import __version__

        msg = "Name:{0} \nApp Version:{1} \n{2} ".format(
            __appname__, __version__, sys.version_info
        )
        QMessageBox.information(self, "Information", msg)

    def showStepsDialog(self):
        msg = stepsInfo(self.lang)
        QMessageBox.information(self, "Information", msg)

    def showKeysDialog(self):
        msg = keysInfo(self.lang)
        QMessageBox.information(self, "Information", msg)

    def createShape(self):
        assert self.beginner()
        self.canvas.setEditing(False)
        self.actions.create.setEnabled(False)
        self.actions.createpoly.setEnabled(False)
        self.canvas.fourpoint = False

    def createPolygon(self):
        assert self.beginner()
        self.canvas.setEditing(False)
        self.canvas.fourpoint = True
        self.actions.create.setEnabled(False)
        self.actions.createpoly.setEnabled(False)
        self.actions.undoLastPoint.setEnabled(True)

    def rotateImg(self, filename, k, _value):
        self.actions.rotateRight.setEnabled(_value)
        pix = cv2.imdecode(np.fromfile(filename, dtype=np.uint8), cv2.IMREAD_COLOR)
        pix = np.rot90(pix, k)
        ext = os.path.splitext(filename)[1]
        cv2.imencode(ext, pix)[1].tofile(filename)
        self.canvas.update()
        self.loadFile(filename)

    def rotateImgWarn(self):
        if self.lang == "ch":
            self.msgBox.warning(
                self,
                "提示",
                "\n 该图片已经有标注框,旋转操作会打乱标注,建议清除标注框后旋转。",
            )
        else:
            self.msgBox.warning(
                self,
                "Warn",
                "\n The picture already has a label box, "
                "and rotation will disrupt the label. "
                "It is recommended to clear the label box and rotate it.",
            )

    def rotateImgAction(self, k=1, _value=False):
        filename = self.filePath

        if os.path.exists(filename):
            if self.itemsToShapesbox:
                self.rotateImgWarn()
            else:
                self.saveFile()
                self.dirty = False
                self.rotateImg(filename=filename, k=k, _value=True)
        else:
            self.rotateImgWarn()
            self.actions.rotateRight.setEnabled(False)
            self.actions.rotateLeft.setEnabled(False)

    def toggleDrawingSensitive(self, drawing=True):
        """In the middle of drawing, toggling between modes should be disabled."""
        self.actions.editMode.setEnabled(not drawing)
        if not drawing and self.beginner():
            # Cancel creation.
            logger.debug("Cancel creation.")
            self.canvas.setEditing(True)
            self.canvas.restoreCursor()
            self.actions.create.setEnabled(True)
            self.actions.createpoly.setEnabled(True)

    def toggleDrawMode(self, edit=True):
        self.canvas.setEditing(edit)
        self.actions.createMode.setEnabled(edit)
        self.actions.editMode.setEnabled(not edit)

    def setCreateMode(self):
        assert self.advanced()
        self.toggleDrawMode(False)

    def setEditMode(self):
        assert self.advanced()
        self.toggleDrawMode(True)
        self.labelSelectionChanged()

    def updateFileMenu(self):
        currFilePath = self.filePath

        def exists(filename):
            return os.path.exists(filename)

        menu = self.menus.recentFiles
        menu.clear()
        files = [f for f in self.recentFiles if f != currFilePath and exists(f)]
        for i, f in enumerate(files):
            icon = newIcon("labels")
            action = QAction(icon, "&%d %s" % (i + 1, QFileInfo(f).fileName()), self)
            action.triggered.connect(partial(self.loadRecent, f))
            menu.addAction(action)

    def popLabelListMenu(self, point):
        self.menus.labelList.exec_(self.labelList.mapToGlobal(point))

    def editLabel(self):
        if not self.canvas.editing():
            return
        item = self.currentItem()
        if not item:
            return
        text = self.labelDialog.popUp(item.text())
        if text is not None:
            item.setText(text)
            # item.setBackground(generateColorByText(text))
            self.setDirty()
            self.updateComboBox()

    # =================== detection box related functions ===================
    def boxItemChanged(self, item):
        shape = self.itemsToShapesbox[item]

        box = ast.literal_eval(item.text())
        # print('shape in labelItemChanged is',shape.points)
        if box != [(int(p.x()), int(p.y())) for p in shape.points]:
            # shape.points = box
            shape.points = [QPointF(p[0], p[1]) for p in box]

            # QPointF(x,y)
            # shape.line_color = generateColorByText(shape.label)
            self.setDirty()
        else:  # User probably changed item visibility
            self.canvas.setShapeVisible(shape, True)  # item.checkState() == Qt.Checked

    def editBox(self):  # ADD
        if not self.canvas.editing():
            return
        item = self.currentBox()
        if not item:
            return
        text = self.labelDialog.popUp(item.text())

        width, height = self.image.width(), self.image.height()
        if text:
            try:
                text_list = eval(text)
            except Exception:
                msg_box = QMessageBox(
                    QMessageBox.Warning, "Warning", "Please enter the correct format"
                )
                msg_box.exec_()
                return
            if len(text_list) < 4:
                msg_box = QMessageBox(
                    QMessageBox.Warning,
                    "Warning",
                    "Please enter the coordinates of 4 points",
                )
                msg_box.exec_()
                return
            for box in text_list:
                if box[0] > width or box[0] < 0 or box[1] > height or box[1] < 0:
                    msg_box = QMessageBox(
                        QMessageBox.Warning, "Warning", "Out of picture size"
                    )
                    msg_box.exec_()
                    return

            item.setText(text)
            # item.setBackground(generateColorByText(text))
            self.setDirty()
            self.updateComboBox()

    def updateBoxlist(self):
        self.canvas.selectedShapes_hShape = []
        if self.canvas.hShape is not None:
            self.canvas.selectedShapes_hShape = self.canvas.selectedShapes + [
                self.canvas.hShape
            ]
        else:
            self.canvas.selectedShapes_hShape = self.canvas.selectedShapes
        for shape in self.canvas.selectedShapes_hShape:
            if shape in self.shapesToItemsbox.keys():
                item = self.shapesToItemsbox[shape]  # listitem
                text = [(int(p.x()), int(p.y())) for p in shape.points]
                item.setText(str(text))
        self.actions.undo.setEnabled(True)
        self.setDirty()

    def indexTo5Files(self, currIndex):
        if currIndex < 2:
            return self.mImgList[:5]
        elif currIndex > len(self.mImgList) - 3:
            return self.mImgList[-5:]
        else:
            return self.mImgList[currIndex - 2 : currIndex + 3]

    # Tzutalin 20160906 : Add file list and dock to move faster
    def fileitemDoubleClicked(self, item=None):
        self.currIndex = self.mImgList.index(
            os.path.join(os.path.abspath(self.dirname), item.text())
        )
        filename = self.mImgList[self.currIndex]
        if filename:
            self.mImgList5 = self.indexTo5Files(self.currIndex)
            # self.additems5(None)
            self.loadFile(filename)
        if self.autoImportOption.isChecked():
            self.importhtml()

    def iconitemDoubleClicked(self, item=None):
        self.currIndex = self.mImgList.index(os.path.join(item.toolTip()))
        filename = self.mImgList[self.currIndex]
        if filename:
            self.mImgList5 = self.indexTo5Files(self.currIndex)
            # self.additems5(None)
            self.loadFile(filename)
        if self.autoImportOption.isChecked():
            self.importhtml()

    def CanvasSizeChange(self):
        if len(self.mImgList) > 0 and self.imageSlider.hasFocus():
            self.zoomWidget.setValue(self.imageSlider.value())

    def shapeSelectionChanged(self, selected_shapes):
        self._noSelectionSlot = True
        for shape in self.canvas.selectedShapes:
            shape.selected = False
        self.labelList.clearSelection()
        self.indexList.clearSelection()
        self.canvas.selectedShapes = selected_shapes
        for shape in self.canvas.selectedShapes:
            shape.selected = True
            self.shapesToItems[shape].setSelected(True)
            self.shapesToItemsbox[shape].setSelected(True)
            index = self.labelList.indexFromItem(self.shapesToItems[shape]).row()
            self.indexList.item(index).setSelected(True)

        self.labelList.scrollToItem(
            self.currentItem()
        )  # QAbstractItemView.EnsureVisible
        # map current label item to index item and select it
        index = self.labelList.indexFromItem(self.currentItem()).row()
        self.indexList.scrollToItem(self.indexList.item(index))
        self.BoxList.scrollToItem(self.currentBox())

        if self.kie_mode:
            if len(self.canvas.selectedShapes) == 1 and self.keyList.count() > 0:
                selected_key_item_row = self.keyList.findItemsByLabel(
                    self.canvas.selectedShapes[0].key_cls, get_row=True
                )
                if (
                    isinstance(selected_key_item_row, list)
                    and len(selected_key_item_row) == 0
                ):
                    key_text = self.canvas.selectedShapes[0].key_cls
                    item = self.keyList.createItemFromLabel(key_text)
                    self.keyList.addItem(item)
                    rgb = self._get_rgb_by_label(key_text, self.kie_mode)
                    self.keyList.setItemLabel(item, key_text, rgb)
                    selected_key_item_row = self.keyList.findItemsByLabel(
                        self.canvas.selectedShapes[0].key_cls, get_row=True
                    )

                self.keyList.setCurrentRow(selected_key_item_row)

        self._noSelectionSlot = False
        n_selected = len(selected_shapes)
        self.actions.singleRere.setEnabled(n_selected)
        self.actions.cellreRec.setEnabled(n_selected)
        self.actions.delete.setEnabled(n_selected)
        self.actions.copy.setEnabled(n_selected)
        self.actions.edit.setEnabled(n_selected == 1)
        self.actions.lock.setEnabled(n_selected)
        self.actions.change_cls.setEnabled(n_selected)
        self.actions.expand.setEnabled(n_selected)

    def addLabel(self, shape):
        shape.paintLabel = self.displayLabelOption.isChecked()
        shape.paintIdx = self.displayIndexOption.isChecked()

        item = HashableQListWidgetItem(shape.label)
        # current difficult checkbox is disable
        # item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
        # item.setCheckState(Qt.Unchecked) if shape.difficult else item.setCheckState(Qt.Checked)

        # Checked means difficult is False
        # item.setBackground(generateColorByText(shape.label))
        self.itemsToShapes[item] = shape
        self.shapesToItems[shape] = item
        # add current label item index before label string
        current_index = QListWidgetItem(str(self.labelList.count()))
        current_index.setTextAlignment(Qt.AlignHCenter)
        self.indexList.addItem(current_index)
        self.labelList.addItem(item)
        # print('item in add label is ',[(p.x(), p.y()) for p in shape.points], shape.label)

        # ADD for box
        item = HashableQListWidgetItem(
            str([(int(p.x()), int(p.y())) for p in shape.points])
        )
        self.itemsToShapesbox[item] = shape
        self.shapesToItemsbox[shape] = item
        self.BoxList.addItem(item)
        for action in self.actions.onShapesPresent:
            action.setEnabled(True)
        self.updateComboBox()

        # update show counting
        self.BoxListDock.setWindowTitle(
            self.BoxListDockName + f" ({self.BoxList.count()})"
        )
        self.labelListDock.setWindowTitle(
            self.labelListDockName + f" ({self.labelList.count()})"
        )

    def remLabels(self, shapes):
        if shapes is None:
            # print('rm empty label')
            return
        for shape in shapes:
            item = self.shapesToItems[shape]
            self.labelList.takeItem(self.labelList.row(item))
            del self.shapesToItems[shape]
            del self.itemsToShapes[item]
            self.updateComboBox()

            # ADD:
            item = self.shapesToItemsbox[shape]
            self.BoxList.takeItem(self.BoxList.row(item))
            del self.shapesToItemsbox[shape]
            del self.itemsToShapesbox[item]
            self.updateComboBox()
        self.updateIndexList()

    def loadLabels(self, shapes):
        s = []
        shape_index = 0
        for label, points, line_color, key_cls, difficult in shapes:
            shape = Shape(
                label=label,
                line_color=line_color,
                key_cls=key_cls,
                font_family=self.label_font_family,
            )
            for x, y in points:
                # Ensure the labels are within the bounds of the image. If not, fix them.
                x, y, snapped = self.canvas.snapPointToCanvas(x, y)
                if snapped:
                    self.setDirty()

                shape.addPoint(QPointF(x, y))
            shape.difficult = difficult
            shape.idx = shape_index
            shape_index += 1
            # shape.locked = False
            shape.close()
            s.append(shape)

            self._update_shape_color(shape)
            self.addLabel(shape)

        self.updateComboBox()
        self.canvas.loadShapes(s)

    def singleLabel(self, shape):
        if shape is None:
            # print('rm empty label')
            return
        item = self.shapesToItems[shape]
        item.setText(shape.label)
        self.updateComboBox()

        # ADD:
        item = self.shapesToItemsbox[shape]
        item.setText(str([(int(p.x()), int(p.y())) for p in shape.points]))
        self.updateComboBox()

    def updateComboBox(self):
        # Get the unique labels and add them to the Combobox.
        itemsTextList = [
            str(self.labelList.item(i).text()) for i in range(self.labelList.count())
        ]

        uniqueTextList = list(set(itemsTextList))
        # Add a null row for showing all the labels
        uniqueTextList.append("")
        uniqueTextList.sort()

        # self.comboBox.update_items(uniqueTextList)

    def updateIndexList(self):
        self.indexList.clear()
        for i in range(self.labelList.count()):
            string = QListWidgetItem(str(i))
            string.setTextAlignment(Qt.AlignHCenter)
            self.indexList.addItem(string)

    def saveLabels(self, annotationFilePath, mode="Auto"):
        # Mode is Auto means that labels will be loaded from self.result_dic totally, which is the output of ocr model
        annotationFilePath = annotationFilePath

        def format_shape(s):
            # print('s in saveLabels is ',s)
            return dict(
                label=s.label,  # str
                line_color=s.line_color.getRgb(),
                fill_color=s.fill_color.getRgb(),
                points=[(int(p.x()), int(p.y())) for p in s.points],  # QPonitF
                difficult=s.difficult,
                key_cls=s.key_cls,
            )  # bool

        if mode == "Auto":
            shapes = []
        else:
            shapes = [
                format_shape(shape)
                for shape in self.canvas.shapes
                if shape.line_color != DEFAULT_LOCK_COLOR
            ]
        # Can add different annotation formats here
        for box in self.result_dic:
            trans_dic = {"label": box[1][0], "points": box[0], "difficult": False}
            if self.kie_mode:
                if len(box) == 3:
                    trans_dic.update({"key_cls": box[2]})
                else:
                    trans_dic.update({"key_cls": "None"})
            if trans_dic["label"] == "" and mode == "Auto":
                continue
            shapes.append(trans_dic)

        try:
            trans_dic = []
            for box in shapes:
                trans_dict = {
                    "transcription": box["label"],
                    "points": box["points"],
                    "difficult": box["difficult"],
                }
                if self.kie_mode:
                    trans_dict.update({"key_cls": box["key_cls"]})
                trans_dic.append(trans_dict)
            self.PPlabel[annotationFilePath] = trans_dic
            if mode == "Auto":
                self.Cachelabel[annotationFilePath] = trans_dic

            # else:
            #     self.labelFile.save(annotationFilePath, shapes, self.filePath, self.imageData,
            #                         self.lineColor.getRgb(), self.fillColor.getRgb())
            # print('Image:{0} -> Annotation:{1}'.format(self.filePath, annotationFilePath))
            return True
        except Exception:
            self.errorMessage("Error saving label data", "Error saving label data")
            return False

    def copySelectedShape(self):
        for shape in self.canvas.copySelectedShape():
            self.addLabel(shape)
        # fix copy and delete
        # self.shapeSelectionChanged(True)

    def move_scrollbar(self, value):
        self.labelListBar.setValue(int(value))
        self.indexListBar.setValue(int(value))

    def labelSelectionChanged(self):
        if self._noSelectionSlot:
            return
        if self.canvas.editing():
            selected_shapes = []
            for item in self.labelList.selectedItems():
                selected_shapes.append(self.itemsToShapes[item])
            if selected_shapes:
                self.canvas.selectShapes(selected_shapes)
            else:
                self.canvas.deSelectShape()

    def indexSelectionChanged(self):
        if self._noSelectionSlot:
            return
        if self.canvas.editing():
            selected_shapes = []
            for item in self.indexList.selectedItems():
                # map index item to label item
                index = self.indexList.indexFromItem(item).row()
                item = self.labelList.item(index)
                selected_shapes.append(self.itemsToShapes[item])
            if selected_shapes:
                self.canvas.selectShapes(selected_shapes)
            else:
                self.canvas.deSelectShape()

    def boxSelectionChanged(self):
        if self._noSelectionSlot:
            # self.BoxList.scrollToItem(self.currentBox(), QAbstractItemView.PositionAtCenter)
            return
        if self.canvas.editing():
            selected_shapes = []
            for item in self.BoxList.selectedItems():
                selected_shapes.append(self.itemsToShapesbox[item])
            if selected_shapes:
                self.canvas.selectShapes(selected_shapes)
            else:
                self.canvas.deSelectShape()

    def labelItemChanged(self, item):
        # avoid accidentally triggering the itemChanged siganl with unhashable item
        # Unknown trigger condition
        if isinstance(item, HashableQListWidgetItem):
            shape = self.itemsToShapes[item]
            label = item.text()
            if label != shape.label:
                shape.label = item.text()
                # shape.line_color = generateColorByText(shape.label)
                self.setDirty()
            elif not ((item.checkState() == Qt.Unchecked) ^ (not shape.difficult)):
                shape.difficult = True if item.checkState() == Qt.Unchecked else False
                self.setDirty()
            else:  # User probably changed item visibility
                self.canvas.setShapeVisible(
                    shape, True
                )  # item.checkState() == Qt.Checked
                # self.actions.save.setEnabled(True)
        else:
            logger.warning(
                "enter labelItemChanged slot with unhashable item: %s %s",
                item,
                item.text(),
            )

    def drag_drop_happened(self):
        """
        label list drag drop signal slot
        """
        # should only select single item
        for item in self.labelList.selectedItems():
            newIndex = self.labelList.indexFromItem(item).row()

        # only support drag_drop one item
        assert len(self.canvas.selectedShapes) > 0
        for shape in self.canvas.selectedShapes:
            selectedShapeIndex = shape.idx

        if newIndex == selectedShapeIndex:
            return

        # move corresponding item in shape list
        shape = self.canvas.shapes.pop(selectedShapeIndex)
        self.canvas.shapes.insert(newIndex, shape)

        # update bbox index
        self.canvas.updateShapeIndex()

        # boxList update simultaneously
        item = self.BoxList.takeItem(selectedShapeIndex)
        self.BoxList.insertItem(newIndex, item)

        # changes happen
        self.setDirty()

    # Callback functions:
    def newShape(self, value=True):
        """Pop-up and give focus to the label editor.

        position MUST be in global coordinates.
        """
        if len(self.labelHist) > 0:
            self.labelDialog = LabelDialog(parent=self, listItem=self.labelHist)

        if value:
            text = self.labelDialog.popUp(text=self.prevLabelText)
            self.lastLabel = text
        else:
            text = self.prevLabelText

        if text is not None:
            self.prevLabelText = self.stringBundle.getString("tempLabel")

            shape = self.canvas.setLastLabel(
                text, None, None, None
            )  # generate_color, generate_color
            if self.kie_mode:
                key_text, _ = self.keyDialog.popUp(self.key_previous_text)
                if key_text is not None:
                    shape = self.canvas.setLastLabel(
                        text, None, None, key_text
                    )  # generate_color, generate_color
                    self.key_previous_text = key_text
                    if not self.keyList.findItemsByLabel(key_text):
                        item = self.keyList.createItemFromLabel(key_text)
                        self.keyList.addItem(item)
                        rgb = self._get_rgb_by_label(key_text, self.kie_mode)
                        self.keyList.setItemLabel(item, key_text, rgb)

                    self._update_shape_color(shape)
                    self.keyDialog.addLabelHistory(key_text)

            self.addLabel(shape)
            if self.beginner():  # Switch to edit mode.
                self.canvas.setEditing(True)
                self.actions.create.setEnabled(True)
                self.actions.createpoly.setEnabled(True)
                self.actions.undoLastPoint.setEnabled(False)
                self.actions.undo.setEnabled(True)
            else:
                self.actions.editMode.setEnabled(True)
            self.setDirty()

            if self.autoReRecognitionOption.isChecked():
                self.reRecognition()
        else:
            # self.canvas.undoLastLine()
            self.canvas.resetAllLines()

    def _update_shape_color(self, shape):
        r, g, b = self._get_rgb_by_label(shape.key_cls, self.kie_mode)
        shape.line_color = QColor(r, g, b)
        shape.vertex_fill_color = QColor(r, g, b)
        shape.hvertex_fill_color = QColor(255, 255, 255)
        shape.fill_color = QColor(r, g, b, 32)
        shape.select_line_color = QColor(
            self.selected_shape_color[0],
            self.selected_shape_color[1],
            self.selected_shape_color[2],
        )
        shape.select_fill_color = QColor(r, g, b, 32)

    def _get_rgb_by_label(self, label, kie_mode):
        shift_auto_shape_color = 2  # use for random color
        if kie_mode and label != "None":
            item = self.keyList.findItemsByLabel(label)[0]
            label_id = self.keyList.indexFromItem(item).row() + 1
            label_id += shift_auto_shape_color
            return LABEL_COLORMAP[label_id % len(LABEL_COLORMAP)]
        else:
            return 0, 255, 0

    def scrollRequest(self, delta, orientation):
        units = -delta / (8 * 15)
        bar = self.scrollBars[orientation]
        bar.setValue(int(bar.value() + bar.singleStep() * units))

    def setZoom(self, value):
        self.actions.fitWidth.setChecked(False)
        self.actions.fitWindow.setChecked(False)
        self.zoomMode = self.MANUAL_ZOOM
        self.zoomWidget.setValue(int(value))

    def addZoom(self, increment=10):
        self.setZoom(int(self.zoomWidget.value() + increment))
        self.imageSlider.setValue(
            int(self.zoomWidget.value() + increment)
        )  # set zoom slider value

    def zoomRequest(self, delta, pos: QPoint = None):
        # get the current scrollbar positions
        # calculate the percentages ~ coordinates
        h_bar = self.scrollBars[Qt.Horizontal]
        v_bar = self.scrollBars[Qt.Vertical]

        # get the current maximum, to know the difference after zooming
        h_bar_max = h_bar.maximum()
        v_bar_max = v_bar.maximum()

        # get the cursor position and canvas size
        # calculate the desired movement from 0 to 1
        # where 0 = move left
        #       1 = move right
        # up and down analogous
        if pos is None:
            cursor = QCursor()
            pos = cursor.pos()

        relative_pos = QWidget.mapFromGlobal(self, pos)

        cursor_x = relative_pos.x()
        cursor_y = relative_pos.y()

        w = self.scrollArea.width()
        h = self.scrollArea.height()

        # the scaling from 0 to 1 has some padding
        # you don't have to hit the very leftmost pixel for a maximum-left movement
        margin = 0.1
        move_x = (cursor_x - margin * w) / (w - 2 * margin * w)
        move_y = (cursor_y - margin * h) / (h - 2 * margin * h)

        # clamp the values from 0 to 1
        move_x = min(max(move_x, 0), 1)
        move_y = min(max(move_y, 0), 1)

        # zoom in
        units = delta / (8 * 15)
        scale = 10
        self.addZoom(scale * units)

        # get the difference in scrollbar values
        # this is how far we can move
        d_h_bar_max = h_bar.maximum() - h_bar_max
        d_v_bar_max = v_bar.maximum() - v_bar_max

        # get the new scrollbar values
        new_h_bar_value = h_bar.value() + move_x * d_h_bar_max
        new_v_bar_value = v_bar.value() + move_y * d_v_bar_max

        h_bar.setValue(int(new_h_bar_value))
        v_bar.setValue(int(new_v_bar_value))

    def setFitWindow(self, value=True):
        if value:
            self.actions.fitWidth.setChecked(False)
        self.zoomMode = self.FIT_WINDOW if value else self.MANUAL_ZOOM
        self.adjustScale()

    def setFitWidth(self, value=True):
        if value:
            self.actions.fitWindow.setChecked(False)
        self.zoomMode = self.FIT_WIDTH if value else self.MANUAL_ZOOM
        self.adjustScale()

    def togglePolygons(self, value):
        for item, shape in self.itemsToShapes.items():
            self.canvas.setShapeVisible(shape, value)

    def loadFile(self, filePath=None, isAdjustScale=True):
        """Load the specified file, or the last opened file if None."""
        self.canvas.shape_move_index = None
        if self.dirty:
            self.mayContinue()
        self.resetState()
        self.canvas.setEnabled(False)
        if filePath is None:
            filePath = self.settings.get(SETTING_FILENAME)

        # Make sure that filePath is a regular python string, rather than QString
        filePath = filePath
        # Fix bug: An index error after select a directory when open a new file.
        unicodeFilePath = filePath
        # unicodeFilePath = os.path.abspath(unicodeFilePath)
        # Tzutalin 20160906 : Add file list and dock to move faster
        # Highlight the file item

        if unicodeFilePath and self.fileListWidget.count() > 0:
            if unicodeFilePath in self.mImgList:
                index = self.mImgList.index(unicodeFilePath)
                fileWidgetItem = self.fileListWidget.item(index)
                logger.debug("unicodeFilePath is %s", unicodeFilePath)
                fileWidgetItem.setSelected(True)
                self.iconlist.clear()
                self.additems5(None)

                for i in range(5):
                    item_tooltip = self.iconlist.item(i).toolTip()
                    # print(i,"---",item_tooltip)
                    if item_tooltip == filePath:
                        t_item = self.iconlist.item(i)
                        t_item.setSelected(True)
                        self.iconlist.scrollToItem(t_item)
                        break
            else:
                self.fileListWidget.clear()
                self.mImgList.clear()
                self.iconlist.clear()

        # if unicodeFilePath and self.iconList.count() > 0:
        #     if unicodeFilePath in self.mImgList:

        if unicodeFilePath and os.path.exists(unicodeFilePath):
            self.canvas.verified = False
            cvimg = cv2.imdecode(np.fromfile(unicodeFilePath, dtype=np.uint8), 1)
            height, width, depth = cvimg.shape
            cvimg = cv2.cvtColor(cvimg, cv2.COLOR_BGR2RGB)
            image = QImage(
                cvimg.data, width, height, width * depth, QImage.Format_RGB888
            )

            if image.isNull():
                self.errorMessage(
                    "Error opening file",
                    "<p>Make sure <i>%s</i> is a valid image file." % unicodeFilePath,
                )
                self.status("Error reading %s" % unicodeFilePath)
                return False
            self.status("Loaded %s" % os.path.basename(unicodeFilePath))
            self.image = image
            self.filePath = unicodeFilePath
            self.canvas.loadPixmap(QPixmap.fromImage(image))

            if self.validFilestate(filePath) is True:
                self.setClean()
            else:
                self.dirty = False
                self.actions.save.setEnabled(True)
            if len(self.canvas.lockedShapes) != 0:
                self.actions.save.setEnabled(True)
                self.setDirty()
            self.canvas.setEnabled(True)
            if isAdjustScale:
                self.adjustScale(initial=True)
            self.paintCanvas()
            self.addRecentFile(self.filePath)
            self.toggleActions(True)

            self.showBoundingBoxFromPPlabel(filePath)

            self.setWindowTitle(__appname__ + " " + filePath)

            # Default : select last item if there is at least one item
            if self.labelList.count():
                self.labelList.setCurrentItem(
                    self.labelList.item(self.labelList.count() - 1)
                )
                self.labelList.item(self.labelList.count() - 1).setSelected(True)
                self.indexList.item(self.labelList.count() - 1).setSelected(True)

            # show file list image count
            select_indexes = self.fileListWidget.selectedIndexes()
            if len(select_indexes) > 0:
                self.fileDock.setWindowTitle(
                    self.fileListName + f" ({select_indexes[0].row() + 1}"
                    f"/{self.fileListWidget.count()})"
                )
            # update show counting
            self.BoxListDock.setWindowTitle(
                self.BoxListDockName + f" ({self.BoxList.count()})"
            )
            self.labelListDock.setWindowTitle(
                self.labelListDockName + f" ({self.labelList.count()})"
            )

            self.canvas.setFocus(True)

            if self.bbox_auto_zoom_center:
                if len(self.canvas.shapes) > 0:
                    (
                        center_x,
                        center_y,
                        shape_area,
                    ) = polygon_bounding_box_center_and_area(
                        self.canvas.shapes[0].points
                    )
                    if shape_area < 30000:
                        zoom_value = 120 * map_value(shape_area, 100, 30000, 20, 0)
                        self.zoomRequest(zoom_value, QPoint(center_x, center_y))
                        # print(" =========> ", shape_area, " ==> ", zoom_value)
            return True
        return False

    def showBoundingBoxFromPPlabel(self, filePath):
        width, height = self.image.width(), self.image.height()
        img_idx = self.getImglabelidx(filePath)
        shapes = []
        # box['ratio'] of the shapes saved in lockedShapes contains the ratio of the
        # four corner coordinates of the shapes to the height and width of the image
        for box in self.canvas.lockedShapes:
            key_cls = "None" if not self.kie_mode else box["key_cls"]
            if self.canvas.isInTheSameImage:
                shapes.append(
                    (
                        box["transcription"],
                        [[s[0] * width, s[1] * height] for s in box["ratio"]],
                        DEFAULT_LOCK_COLOR,
                        key_cls,
                        box["difficult"],
                    )
                )
            else:
                shapes.append(
                    (
                        "锁定框：待检测",
                        [[s[0] * width, s[1] * height] for s in box["ratio"]],
                        DEFAULT_LOCK_COLOR,
                        key_cls,
                        box["difficult"],
                    )
                )
        if img_idx in self.PPlabel.keys():
            for box in self.PPlabel[img_idx]:
                key_cls = "None" if not self.kie_mode else box.get("key_cls", "None")
                shapes.append(
                    (
                        box["transcription"],
                        box["points"],
                        None,
                        key_cls,
                        box.get("difficult", False),
                    )
                )

        if shapes:
            self.loadLabels(shapes)
            self.canvas.verified = False

    def validFilestate(self, filePath):
        if filePath in self.fileStatedict.keys() and self.fileStatedict[filePath] == 1:
            return True
        elif (
            self.getImglabelidx(filePath) in self.fileStatedict.keys()
            and self.fileStatedict[self.getImglabelidx(filePath)] == 1
        ):
            return True
        else:
            return False

    def resizeEvent(self, event):
        if (
            self.canvas
            and not self.image.isNull()
            and self.zoomMode != self.MANUAL_ZOOM
        ):
            self.adjustScale()
        super(MainWindow, self).resizeEvent(event)

    def paintCanvas(self):
        assert not self.image.isNull(), "cannot paint null image"
        self.canvas.scale = 0.01 * self.zoomWidget.value()
        self.canvas.adjustSize()
        self.canvas.update()

    def adjustScale(self, initial=False):
        value = self.scalers[self.FIT_WINDOW if initial else self.zoomMode]()
        self.zoomWidget.setValue(int(100 * value))
        self.imageSlider.setValue(self.zoomWidget.value())  # set zoom slider value

    def scaleFitWindow(self):
        """Figure out the size of the pixmap in order to fit the main widget."""
        e = 2.0  # So that no scrollbars are generated.
        w1 = self.centralWidget().width() - e
        h1 = self.centralWidget().height() - e - 110
        a1 = w1 / h1
        # Calculate a new scale value based on the pixmap's aspect ratio.
        w2 = self.canvas.pixmap.width() - 0.0
        h2 = self.canvas.pixmap.height() - 0.0
        a2 = w2 / h2
        return w1 / w2 if a2 >= a1 else h1 / h2

    def scaleFitWidth(self):
        # The epsilon does not seem to work too well here.
        w = self.centralWidget().width() - 2.0
        return w / self.canvas.pixmap.width()

    def closeEvent(self, event):
        if not self.mayContinue():
            event.ignore()
        else:
            settings = self.settings
            # If it loads images from dir, don't load it at the beginning
            if self.dirname is None:
                settings[SETTING_FILENAME] = self.filePath if self.filePath else ""
            else:
                settings[SETTING_FILENAME] = ""

            settings[SETTING_WIN_SIZE] = self.size()
            settings[SETTING_WIN_POSE] = self.pos()
            settings[SETTING_WIN_STATE] = self.saveState()
            settings[SETTING_LINE_COLOR] = self.lineColor
            settings[SETTING_FILL_COLOR] = self.fillColor
            settings[SETTING_RECENT_FILES] = self.recentFiles
            settings[SETTING_ADVANCE_MODE] = not self._beginner
            if self.defaultSaveDir and os.path.exists(self.defaultSaveDir):
                settings[SETTING_SAVE_DIR] = self.defaultSaveDir
            else:
                settings[SETTING_SAVE_DIR] = ""

            if self.lastOpenDir and os.path.exists(self.lastOpenDir):
                settings[SETTING_LAST_OPEN_DIR] = self.lastOpenDir
            else:
                settings[SETTING_LAST_OPEN_DIR] = ""

            settings[SETTING_PAINT_LABEL] = self.displayLabelOption.isChecked()
            settings[SETTING_PAINT_INDEX] = self.displayIndexOption.isChecked()
            settings[SETTING_DRAW_SQUARE] = self.drawSquaresOption.isChecked()
            settings.save()
            try:
                self.saveLabelFile()
            except Exception:
                pass

    def loadRecent(self, filename):
        if self.mayContinue():
            logger.info("Loading recent file: %s", filename)
            self.loadFile(filename)

    def scanAllImages(self, folderPath):
        extensions = [
            ".%s" % fmt.data().decode("ascii").lower()
            for fmt in QImageReader.supportedImageFormats()
        ]
        images = []

        for file in os.listdir(folderPath):
            if file.lower().endswith(tuple(extensions)):
                relativePath = os.path.join(folderPath, file)
                path = os.path.abspath(relativePath)
                images.append(path)
        if self.img_list_natural_sort:
            natural_sort(images, key=lambda x: x.lower())
        else:
            images.sort()
        return images

    def openDirDialog(self, _value=False, dirpath=None, silent=False):
        if not self.mayContinue():
            return

        defaultOpenDirPath = dirpath if dirpath else "."
        if self.lastOpenDir and os.path.exists(self.lastOpenDir):
            defaultOpenDirPath = self.lastOpenDir
        else:
            defaultOpenDirPath = (
                os.path.dirname(self.filePath) if self.filePath else "."
            )
        if not silent:
            targetDirPath = QFileDialog.getExistingDirectory(
                self,
                "%s - Open Directory" % __appname__,
                defaultOpenDirPath,
                QFileDialog.ShowDirsOnly | QFileDialog.DontResolveSymlinks,
            )
        else:
            targetDirPath = defaultOpenDirPath
        self.lastOpenDir = targetDirPath
        self.importDirImages(targetDirPath)

    def openDatasetDirDialog(self):
        if self.lastOpenDir and os.path.exists(self.lastOpenDir):
            if platform.system() == "Windows":
                os.startfile(self.lastOpenDir)
            else:
                os.system("open " + os.path.normpath(self.lastOpenDir))
            defaultOpenDirPath = self.lastOpenDir

        else:
            if self.lang == "ch":
                self.msgBox.warning(self, "提示", "\n 原文件夹已不存在,请从新选择数据集路径!")
            else:
                self.msgBox.warning(
                    self,
                    "Warn",
                    "\n The original folder no longer exists, please choose the data set path again!",
                )

            self.actions.open_dataset_dir.setEnabled(False)
            defaultOpenDirPath = (
                os.path.dirname(self.filePath) if self.filePath else "."
            )

    def init_key_list(self, label_dict):
        if not self.kie_mode:
            return
        # load key_cls
        for image, info in label_dict.items():
            for box in info:
                if "key_cls" not in box:
                    box.update({"key_cls": "None"})
                self.existed_key_cls_set.add(box["key_cls"])
        if len(self.existed_key_cls_set) > 0:
            for key_text in self.existed_key_cls_set:
                if not self.keyList.findItemsByLabel(key_text):
                    item = self.keyList.createItemFromLabel(key_text)
                    self.keyList.addItem(item)
                    rgb = self._get_rgb_by_label(key_text, self.kie_mode)
                    self.keyList.setItemLabel(item, key_text, rgb)

        if self.keyDialog is None:
            # key list dialog
            self.keyDialog = KeyDialog(
                text=self.key_dialog_tip,
                parent=self,
                labels=self.existed_key_cls_set,
                sort_labels=True,
                show_text_field=True,
                completion="startswith",
                fit_to_content={"column": True, "row": False},
                flags=None,
            )

    def importDirImages(self, dirpath, isDelete=False):
        if not self.mayContinue() or not dirpath:
            return
        if self.defaultSaveDir and self.defaultSaveDir != dirpath:
            self.saveLabelFile()

        if not isDelete:
            self.loadFilestate(dirpath)
            self.PPlabelpath = dirpath + "/Label.txt"
            self.PPlabel = self.loadLabelFile(self.PPlabelpath)
            self.Cachelabelpath = dirpath + "/Cache.cach"
            self.Cachelabel = self.loadLabelFile(self.Cachelabelpath)
            if self.Cachelabel:
                self.PPlabel = dict(self.Cachelabel, **self.PPlabel)

            self.init_key_list(self.PPlabel)

        self.lastOpenDir = dirpath
        self.dirname = dirpath

        self.defaultSaveDir = dirpath
        self.statusBar().showMessage(
            "%s started. Annotation will be saved to %s"
            % (__appname__, self.defaultSaveDir)
        )
        self.statusBar().show()

        imgListCurrIndex = None
        if self.filePath:
            imgListCurrIndex = self.mImgList.index(self.filePath)

        self.filePath = None
        self.fileListWidget.clear()
        self.mImgList = self.scanAllImages(dirpath)
        self.mImgList5 = self.mImgList[:5]
        self.openNextImg(imgListCurrIndex=imgListCurrIndex)
        doneicon = newIcon("done")
        closeicon = newIcon("close")
        for imgPath in self.mImgList:
            filename = os.path.basename(imgPath)
            if self.validFilestate(imgPath) is True:
                item = QListWidgetItem(doneicon, filename)
            else:
                item = QListWidgetItem(closeicon, filename)
            self.fileListWidget.addItem(item)

        logger.info("DirPath in importDirImages is %s", dirpath)
        self.iconlist.clear()
        self.additems5(dirpath)
        self.changeFileFolder = True
        self.haveAutoReced = False
        self.auto_recognition_num = len(self.mImgList)
        self.AutoRecognitionNum.setRange(0, len(self.mImgList))
        self.AutoRecognitionNum.setValue(self.auto_recognition_num)
        self.AutoRecognition.setEnabled(True)
        self.reRecogButton.setEnabled(True)
        self.tableRecButton.setEnabled(True)
        self.actions.AutoRec.setEnabled(True)
        self.actions.reRec.setEnabled(True)
        self.actions.tableRec.setEnabled(True)
        self.actions.open_dataset_dir.setEnabled(True)
        self.actions.rotateLeft.setEnabled(True)
        self.actions.rotateRight.setEnabled(True)

        fileListWidgetCurrentRow = 0
        if imgListCurrIndex is not None:
            fileListWidgetCurrentRow = imgListCurrIndex
            if fileListWidgetCurrentRow >= self.fileListWidget.count():
                fileListWidgetCurrentRow = fileListWidgetCurrentRow - 1

        self.fileListWidget.setCurrentRow(
            fileListWidgetCurrentRow
        )  # set list index to first
        self.fileDock.setWindowTitle(
            self.fileListName
            + f" ({fileListWidgetCurrentRow + 1}/{self.fileListWidget.count()})"
        )  # show image count

    def openPrevImg(self, _value=False):
        if len(self.mImgList) <= 0:
            return

        if self.filePath is None:
            return

        currIndex = self.mImgList.index(self.filePath)
        self.mImgList5 = self.mImgList[:5]
        if currIndex - 1 >= 0:
            self.currIndex = self.currIndex - 1
            filename = self.mImgList[currIndex - 1]
            self.mImgList5 = self.indexTo5Files(currIndex - 1)
            if filename:
                self.loadFile(filename)
        if self.autoImportOption.isChecked():
            self.importhtml()

    def openNextImg(self, _value=False, imgListCurrIndex=None):
        if not self.mayContinue():
            return

        if len(self.mImgList) <= 0:
            return

        filename = None
        if self.filePath is None and imgListCurrIndex is None:
            filename = self.mImgList[0]
            self.mImgList5 = self.mImgList[:5]
        else:
            if imgListCurrIndex is None:
                currIndex = self.mImgList.index(self.filePath)
            else:
                currIndex = imgListCurrIndex - 1

            if currIndex + 1 < len(self.mImgList):
                self.currIndex = self.currIndex + 1
                filename = self.mImgList[currIndex + 1]
                self.mImgList5 = self.indexTo5Files(currIndex + 1)
            else:
                filename = self.mImgList[currIndex]
                self.mImgList5 = self.indexTo5Files(currIndex)
        if filename:
            logger.debug("file name in openNext is %s", filename)
            self.loadFile(filename)
        if self.autoImportOption.isChecked():
            self.importhtml()

    def updateFileListIcon(self, filename):
        pass

    def saveFile(self, _value=False, mode="Manual"):
        # Manual mode is used for users click "Save" manually,which will change the state of the image
        if self.filePath:
            img_idx = self.getImglabelidx(self.filePath)
            self._saveFile(img_idx, mode=mode)

    def saveLockedShapes(self):
        self.canvas.lockedShapes = []
        self.canvas.selectedShapes = []
        for s in self.canvas.shapes:
            if s.line_color == DEFAULT_LOCK_COLOR:
                self.canvas.selectedShapes.append(s)
        self.lockSelectedShape()
        for s in self.canvas.shapes:
            if s.line_color == DEFAULT_LOCK_COLOR:
                self.canvas.selectedShapes.remove(s)
                self.canvas.shapes.remove(s)

    def _saveFile(self, annotationFilePath, mode="Manual"):
        if len(self.canvas.lockedShapes) != 0:
            self.saveLockedShapes()

        if mode == "Manual":
            self.result_dic_locked = []
            img = cv2.imdecode(
                np.fromfile(self.filePath, dtype=np.uint8), cv2.IMREAD_COLOR
            )
            width, height = self.image.width(), self.image.height()
            for shape in self.canvas.lockedShapes:
                box = [[int(p[0] * width), int(p[1] * height)] for p in shape["ratio"]]
                # assert len(box) == 4
                result = [(shape["transcription"], 1)]
                result.insert(0, box)
                self.result_dic_locked.append(result)
            self.result_dic += self.result_dic_locked
            self.result_dic_locked = []
            if annotationFilePath and self.saveLabels(annotationFilePath, mode=mode):
                self.setClean()
                self.statusBar().showMessage("Saved to  %s" % annotationFilePath)
                self.statusBar().show()
                currIndex = self.mImgList.index(self.filePath)
                item = self.fileListWidget.item(currIndex)
                item.setIcon(newIcon("done"))

                self.fileStatedict[self.getImglabelidx(self.filePath)] = 1
                if len(self.fileStatedict) % self.autoSaveNum == 0:
                    self.saveFilestate()
                    self.savePPlabel(mode="Auto")

                self.fileListWidget.insertItem(int(currIndex), item)
                if not self.canvas.isInTheSameImage:
                    self.openNextImg()
                self.actions.saveRec.setEnabled(True)
                self.actions.saveLabel.setEnabled(True)
                self.actions.exportJSON.setEnabled(True)

        elif mode == "Auto":
            if annotationFilePath and self.saveLabels(annotationFilePath, mode=mode):
                self.setClean()
                self.statusBar().showMessage("Saved to  %s" % annotationFilePath)
                self.statusBar().show()

    def closeFile(self, _value=False):
        if not self.mayContinue():
            return
        self.resetState()
        self.setClean()
        self.toggleActions(False)
        self.canvas.setEnabled(False)
        self.actions.saveAs.setEnabled(False)

    def deleteImg(self):
        deletePath = self.filePath
        if deletePath is not None:
            deleteInfo = self.deleteImgDialog()
            if deleteInfo == QMessageBox.Yes:
                if platform.system() == "Windows":
                    # from win32com import shell, shellcon
                    # shell.SHFileOperation((0, shellcon.FO_DELETE, deletePath, None,
                    #                        shellcon.FOF_SILENT | shellcon.FOF_ALLOWUNDO | shellcon.FOF_NOCONFIRMATION,
                    #                        None, None))
                    os.remove(deletePath)
                    # linux
                elif platform.system() == "Linux":
                    cmd = "trash " + deletePath
                    os.system(cmd)
                    # macOS
                elif platform.system() == "Darwin":
                    import subprocess

                    absPath = (
                        os.path.abspath(deletePath)
                        .replace("\\", "\\\\")
                        .replace('"', '\\"')
                    )
                    cmd = [
                        "osascript",
                        "-e",
                        'tell app "Finder" to move {the POSIX file "'
                        + absPath
                        + '"} to trash',
                    ]
                    logger.debug("Executing command: %s", " ".join(cmd))
                    subprocess.call(cmd, stdout=open(os.devnull, "w"))

                if self.filePath in self.fileStatedict.keys():
                    self.fileStatedict.pop(self.filePath)
                imgidx = self.getImglabelidx(self.filePath)
                if imgidx in self.PPlabel.keys():
                    self.PPlabel.pop(imgidx)

                self.importDirImages(self.lastOpenDir, isDelete=True)

    def deleteImgDialog(self):
        yes, cancel = QMessageBox.Yes, QMessageBox.Cancel
        msg = "The image will be deleted to the recycle bin"
        return QMessageBox.warning(self, "Attention", msg, yes | cancel)

    def resetAll(self):
        self.settings.reset()
        self.close()
        proc = QProcess()
        proc.startDetached(os.path.abspath(__file__))

    def mayContinue(self):  #
        if not self.dirty:
            return True
        else:
            if self.autoSaveUnsavedChangesOption.isChecked():
                self.canvas.isInTheSameImage = True
                self.saveFile()
                self.canvas.isInTheSameImage = False
                return True

            discardChanges = self.discardChangesDialog()
            if discardChanges == QMessageBox.No:
                return True
            elif discardChanges == QMessageBox.Yes:
                self.canvas.isInTheSameImage = True
                self.saveFile()
                self.canvas.isInTheSameImage = False
                return True
            else:
                return False

    def discardChangesDialog(self):
        yes, no, cancel = QMessageBox.Yes, QMessageBox.No, QMessageBox.Cancel
        if self.lang == "ch":
            msg = '您有未保存的变更, 您想保存再继续吗?\n点击 "No" 丢弃所有未保存的变更.'
        else:
            msg = 'You have unsaved changes, would you like to save them and proceed?\nClick "No" to undo all changes.'
        return QMessageBox.warning(self, "Attention", msg, yes | no | cancel)

    def errorMessage(self, title, message):
        return QMessageBox.critical(
            self, title, "<p><b>%s</b></p>%s" % (title, message)
        )

    def currentPath(self):
        return os.path.dirname(self.filePath) if self.filePath else "."

    def chooseColor(self):
        color = self.colorDialog.getColor(
            self.lineColor, "Choose line color", default=DEFAULT_LINE_COLOR
        )
        if color:
            self.lineColor = color
            Shape.line_color = color
            self.canvas.setDrawingColor(color)
            self.canvas.update()
            self.setDirty()

    def deleteSelectedShape(self):
        self.remLabels(self.canvas.deleteSelected())
        self.actions.undo.setEnabled(True)
        self.setDirty()
        if self.noShapes():
            for action in self.actions.onShapesPresent:
                action.setEnabled(False)
        self.BoxListDock.setWindowTitle(
            self.BoxListDockName + f" ({self.BoxList.count()})"
        )
        self.labelListDock.setWindowTitle(
            self.labelListDockName + f" ({self.labelList.count()})"
        )

    def chshapeLineColor(self):
        color = self.colorDialog.getColor(
            self.lineColor, "Choose line color", default=DEFAULT_LINE_COLOR
        )
        if color:
            for shape in self.canvas.selectedShapes:
                shape.line_color = color
            self.canvas.update()
            self.setDirty()

    def chshapeFillColor(self):
        color = self.colorDialog.getColor(
            self.fillColor, "Choose fill color", default=DEFAULT_FILL_COLOR
        )
        if color:
            for shape in self.canvas.selectedShapes:
                shape.fill_color = color
            self.canvas.update()
            self.setDirty()

    def copyShape(self):
        self.canvas.endMove(copy=True)
        self.addLabel(self.canvas.selectedShape)
        self.setDirty()

    def moveShape(self):
        self.canvas.endMove(copy=False)
        self.setDirty()

    def loadPredefinedClasses(self, predefClassesFile):
        if os.path.exists(predefClassesFile) is True:
            with codecs.open(predefClassesFile, "r", "utf8") as f:
                for line in f:
                    line = line.strip()
                    if self.labelHist is None:
                        self.labelHist = [line]
                    else:
                        self.labelHist.append(line)

    def togglePaintLabelsOption(self):
        self.displayIndexOption.setChecked(False)
        for shape in self.canvas.shapes:
            shape.paintLabel = self.displayLabelOption.isChecked()
            shape.paintIdx = self.displayIndexOption.isChecked()
        self.canvas.repaint()

    def togglePaintIndexOption(self):
        self.displayLabelOption.setChecked(False)
        for shape in self.canvas.shapes:
            shape.paintLabel = self.displayLabelOption.isChecked()
            shape.paintIdx = self.displayIndexOption.isChecked()
        self.canvas.repaint()

    def toogleDrawSquare(self):
        self.canvas.setDrawingShapeToSquare(self.drawSquaresOption.isChecked())

    def additems(self, dirpath):
        for file in self.mImgList:
            pix = QPixmap(file)
            _, filename = os.path.split(file)
            filename, _ = os.path.splitext(filename)
            item = QListWidgetItem(
                QIcon(
                    pix.scaled(100, 100, Qt.IgnoreAspectRatio, Qt.FastTransformation)
                ),
                filename[:10],
            )
            item.setToolTip(file)
            self.iconlist.addItem(item)

    def additems5(self, dirpath):
        for file in self.mImgList5:
            pix = QPixmap(file)
            _, filename = os.path.split(file)
            filename, _ = os.path.splitext(filename)
            pfilename = filename[:10]
            if len(pfilename) < 10:
                lentoken = 12 - len(pfilename)
                prelen = lentoken // 2
                bfilename = prelen * " " + pfilename + (lentoken - prelen) * " "
            # item = QListWidgetItem(QIcon(pix.scaled(100, 100, Qt.KeepAspectRatio, Qt.SmoothTransformation)),filename[:10])
            item = QListWidgetItem(
                QIcon(
                    pix.scaled(100, 100, Qt.IgnoreAspectRatio, Qt.FastTransformation)
                ),
                pfilename,
            )
            # item.setForeground(QBrush(Qt.white))
            item.setToolTip(file)
            self.iconlist.addItem(item)
        owidth = 0
        for index in range(len(self.mImgList5)):
            item = self.iconlist.item(index)
            itemwidget = self.iconlist.visualItemRect(item)
            owidth += itemwidget.width()
        self.iconlist.setMinimumWidth(owidth + 50)

    def gen_quad_from_poly(self, poly):
        """
        Generate min area quad from poly.
        """
        point_num = poly.shape[0]
        min_area_quad = np.zeros((4, 2), dtype=np.float32)
        rect = cv2.minAreaRect(
            poly.astype(np.int32)
        )  # (center (x,y), (width, height), angle of rotation)
        box = np.array(cv2.boxPoints(rect))

        first_point_idx = 0
        min_dist = 1e4
        for i in range(4):
            dist = (
                np.linalg.norm(box[(i + 0) % 4] - poly[0])
                + np.linalg.norm(box[(i + 1) % 4] - poly[point_num // 2 - 1])
                + np.linalg.norm(box[(i + 2) % 4] - poly[point_num // 2])
                + np.linalg.norm(box[(i + 3) % 4] - poly[-1])
            )
            if dist < min_dist:
                min_dist = dist
                first_point_idx = i
        for i in range(4):
            min_area_quad[i] = box[(first_point_idx + i) % 4]

        bbox_new = min_area_quad.tolist()
        bbox = []

        for box in bbox_new:
            box = list(map(int, box))
            bbox.append(box)

        return bbox

    def getImglabelidx(self, filePath):
        if platform.system() == "Windows":
            spliter = "\\"
        else:
            spliter = "/"
        file_path_split = filePath.split(spliter)[-2:]
        if len(file_path_split) == 1:
            return filePath
        return file_path_split[0] + "/" + file_path_split[1]

    def autoRecognitionNum(self, value):
        remain_num = len(self.mImgList) - self.currIndex
        if value == 0:
            self.auto_recognition_num = remain_num
        else:
            self.auto_recognition_num = min(value, remain_num)
        self.AutoRecognitionNum.setValue(self.auto_recognition_num)

    def autoRecognition(self):
        assert self.mImgList is not None
        logger.info("Using model from %s", self.model)

        start_index = self.currIndex
        end_index = min(self.currIndex + self.auto_recognition_num, len(self.mImgList))
        images_to_check = self.mImgList[start_index:end_index]

        recorded_basenames = [
            os.path.basename(path)
            for path in self.fileStatedict.keys()
            if self.fileStatedict[path] == 1
        ]

        uncheckedList = []
        for image_path in images_to_check:
            image_basename = os.path.basename(image_path)
            if image_basename not in recorded_basenames:
                uncheckedList.append(image_path)

        self.autoDialog = AutoDialog(
            parent=self,
            ocr=self.ocr,
            image_list=uncheckedList,
            len_bar=len(uncheckedList),
        )
        self.autoDialog.popUp()
        self.haveAutoReced = True
        self.filePath = self.mImgList[self.currIndex]
        self.loadFile(self.filePath, isAdjustScale=False)
        self.saveCacheLabel()

        self.init_key_list(self.Cachelabel)

    def reRecognition(self):
        img = cv2.imdecode(np.fromfile(self.filePath, dtype=np.uint8), 1)
        if self.canvas.shapes:
            self.result_dic = []
            self.result_dic_locked = (
                []
            )  # result_dic_locked stores the ocr result of self.canvas.lockedShapes
            rec_flag = 0
            for shape in self.canvas.shapes:
                box = [[int(p.x()), int(p.y())] for p in shape.points]
                kie_cls = shape.key_cls

                if len(box) > 4:
                    box = self.gen_quad_from_poly(np.array(box))
                assert len(box) == 4

                img_crop = get_rotate_crop_image(img, np.array(box, np.float32))
                if img_crop is None:
                    msg = (
                        "Can not recognise the detection box in "
                        + self.filePath
                        + ". Please change manually"
                    )
                    QMessageBox.information(self, "Information", msg)
                    return
                result = self.text_recognizer.predict(img_crop)[0]
                storage = [(result["rec_text"], result["rec_score"])]
                if result["rec_text"] != "":
                    if shape.line_color == DEFAULT_LOCK_COLOR:
                        shape.label = result["rec_text"]
                        storage.insert(0, box)
                        if self.kie_mode:
                            storage.append(kie_cls)
                        self.result_dic_locked.append(storage)
                    else:
                        storage.insert(0, box)
                        if self.kie_mode:
                            storage.append(kie_cls)
                        self.result_dic.append(storage)
                else:
                    logger.warning("Can not recognise the box")
                    if shape.line_color == DEFAULT_LOCK_COLOR:
                        shape.label = result["rec_text"]
                        if self.kie_mode:
                            self.result_dic_locked.append(
                                [box, (self.noLabelText, 0), kie_cls]
                            )
                        else:
                            self.result_dic_locked.append([box, (self.noLabelText, 0)])
                    else:
                        if self.kie_mode:
                            self.result_dic.append(
                                [box, (self.noLabelText, 0), kie_cls]
                            )
                        else:
                            self.result_dic.append([box, (self.noLabelText, 0)])
                try:
                    if (
                        self.noLabelText == shape.label
                        or result["rec_text"] == shape.label
                    ):
                        logger.debug("label no change")
                    else:
                        rec_flag += 1
                except IndexError as e:
                    logger.warning("Can not recognise the box")
            if (len(self.result_dic) > 0 and rec_flag > 0) or self.canvas.lockedShapes:
                self.canvas.isInTheSameImage = True
                self.saveFile(mode="Auto")
                self.loadFile(self.filePath, isAdjustScale=False)
                self.canvas.isInTheSameImage = False
                self.setDirty()
            elif len(self.result_dic) == len(self.canvas.shapes) and rec_flag == 0:
                if self.lang == "ch":
                    QMessageBox.information(self, "Information", "识别结果保持一致！")
                else:
                    QMessageBox.information(
                        self, "Information", "The recognition result remains unchanged!"
                    )
            else:
                logger.warning("Can not recognise in %s", self.filePath)
        else:
            QMessageBox.information(self, "Information", "Draw a box!")

    def singleRerecognition(self):
        img = cv2.imdecode(np.fromfile(self.filePath, dtype=np.uint8), 1)
        for shape in self.canvas.selectedShapes:
            box = [[int(p.x()), int(p.y())] for p in shape.points]
            if len(box) > 4:
                box = self.gen_quad_from_poly(np.array(box))
            assert len(box) == 4
            img_crop = get_rotate_crop_image(img, np.array(box, np.float32))
            if img_crop is None:
                msg = (
                    "Can not recognise the detection box in "
                    + self.filePath
                    + ". Please change manually"
                )
                QMessageBox.information(self, "Information", msg)
                return
            result = self.text_recognizer.predict(img_crop)[0]
            storage = [(result["rec_text"], result["rec_score"])]
            if result["rec_text"] != "":
                storage.insert(0, box)
                storage.append(result["rec_text"])
                if self.kie_mode:
                    storage.append(shape.key_cls)
                logger.debug("result in reRec is %s", result)
                if result["rec_text"] == shape.label:
                    logger.debug("label no change")
                else:
                    shape.label = result["rec_text"]
            else:
                logger.warning("Can not recognise the box")
                if self.noLabelText == shape.label:
                    logger.debug("label no change")
                else:
                    shape.label = self.noLabelText
            self.singleLabel(shape)
            self.setDirty()

    def TableRecognition(self):
        """
        Table Recognition
        """
        from tablepyxl import tablepyxl

        import time

        start = time.time()
        img = cv2.imdecode(np.fromfile(self.filePath, dtype=np.uint8), cv2.IMREAD_COLOR)
        res = self.table_ocr.predict(img)[0]

        table_rec_excel_dir = self.lastOpenDir + "/tableRec_excel_output/"
        os.makedirs(table_rec_excel_dir, exist_ok=True)
        filename, _ = os.path.splitext(os.path.basename(self.filePath))

        excel_path = table_rec_excel_dir + "{}.xlsx".format(filename)

        if res is None:
            msg = (
                "Can not recognise the table in "
                + self.filePath
                + ". Please change manually"
            )
            QMessageBox.information(self, "Information", msg)
            # create an empty excel
            tablepyxl.document_to_xl("", excel_path)
            return

        # save res
        # ONLY SUPPORT ONE TABLE in one image
        has_table_flag = False
        for region in res["table_res_list"]:
            if region["table_ocr_pred"]["rec_boxes"] is None:
                msg = (
                    "Can not recognise the detection box in "
                    + self.filePath
                    + ". Please change manually"
                )
                QMessageBox.information(self, "Information", msg)
                # create an empty excel
                tablepyxl.document_to_xl("", excel_path)
                return
            has_table_flag = True
            # save table ocr result on PPOCRLabel
            # clear all old annotations before saving result
            self.itemsToShapes.clear()
            self.shapesToItems.clear()
            self.itemsToShapesbox.clear()
            self.shapesToItemsbox.clear()
            self.labelList.clear()
            self.indexList.clear()
            self.BoxList.clear()
            self.result_dic = []
            self.result_dic_locked = []

            shapes = []
            result_len = len(region["table_ocr_pred"]["rec_boxes"])
            order_index = 0
            for i in range(result_len):
                bbox = region["table_ocr_pred"]["rec_boxes"][i]
                rec_text = region["table_ocr_pred"]["rec_texts"][i]

                rext_bbox = [
                    [bbox[0], bbox[1]],
                    [bbox[2], bbox[1]],
                    [bbox[2], bbox[3]],
                    [bbox[0], bbox[3]],
                ]

                # save bbox to shape
                shape = Shape(
                    label=rec_text, line_color=DEFAULT_LINE_COLOR, key_cls=None
                )
                for point in rext_bbox:
                    x, y = point
                    # Ensure the labels are within the bounds of the image.
                    # If not, fix them.
                    x, y, _ = self.canvas.snapPointToCanvas(x, y)
                    shape.addPoint(QPointF(x, y))
                shape.difficult = False
                shape.idx = order_index
                order_index += 1
                # shape.locked = False
                shape.close()
                self.addLabel(shape)
                shapes.append(shape)
            self.setDirty()
            self.canvas.loadShapes(shapes)

            # save HTML result to excel
            try:
                tablepyxl.document_to_xl(region["pred_html"], excel_path)
            except Exception as e:
                logger.error(
                    "Can not save excel file. \nError: %s",
                    e,
                )
            break

        if not has_table_flag:
            msg = (
                "Can not recognise the table in "
                + self.filePath
                + ". Please change manually"
            )
            QMessageBox.information(self, "Information", msg)
            # create an empty excel
            try:
                tablepyxl.document_to_xl("", excel_path)
            except AttributeError:  # 如果 tablepyxl 报错，改用 openpyxl
                wb = openpyxl.Workbook()
                wb.save(excel_path)
            return

        # automatically open excel annotation file
        if platform.system() == "Windows":
            try:
                import win32com.client
            except Exception as e:
                logger.error(
                    "CANNOT OPEN .xlsx. It could be one of the following reasons: "
                    "Only support Windows | No python win32com. Error: %s",
                    e,
                )

            try:
                xl = win32com.client.Dispatch("Excel.Application")
                xl.Visible = True
                xl.Workbooks.Open(excel_path)
                # excelEx = "You need to show the excel executable at this point"
                # subprocess.Popen([excelEx, excel_path])

                # os.startfile(excel_path)
            except Exception as e:
                logger.error(
                    "CANNOT OPEN .xlsx. It could be the following reasons: "
                    ".xlsx is not existed. Error: %s",
                    e,
                )
        else:
            os.system("open " + os.path.normpath(excel_path))

        logger.info("Table recognition time cost: %s", time.time() - start)

    def cellreRecognition(self):
        """
        re-recognise text in a cell
        """
        img = cv2.imdecode(np.fromfile(self.filePath, dtype=np.uint8), cv2.IMREAD_COLOR)
        for shape in self.canvas.selectedShapes:
            box = [[int(p.x()), int(p.y())] for p in shape.points]

            if len(box) > 4:
                box = self.gen_quad_from_poly(np.array(box))
            assert len(box) == 4

            # pad around bbox for better text recognition accuracy
            _box = boxPad(box, img.shape, 6)
            img_crop = get_rotate_crop_image(img, np.array(_box, np.float32))
            if img_crop is None:
                msg = (
                    "Can not recognise the detection box in "
                    + self.filePath
                    + ". Please change manually"
                )
                QMessageBox.information(self, "Information", msg)
                return

            # merge the text result in the cell
            texts = ""
            probs = 0.0  # the probability of the cell is average prob of every text box in the cell
            det_res = self.text_detector.predict(img_crop)[0]
            bboxes = det_res["dt_polys"].tolist()
            if len(bboxes) > 0:
                bboxes.reverse()  # top row text at first
                for _bbox in bboxes:
                    patch = get_rotate_crop_image(img_crop, np.array(_bbox, np.float32))
                    rec_res = self.text_recognizer.predict(patch)[0]
                    text = rec_res["rec_text"]
                    if text != "":
                        texts += text + (
                            "" if text[0].isalpha() else " "
                        )  # add space between english word
                        probs += rec_res["rec_score"]
                probs = probs / len(bboxes)
            result = [(texts.strip(), probs)]

            if result[0][0] != "":
                result.insert(0, box)
                logger.debug("result in reRec is %s", result)
                if result[1][0] == shape.label:
                    logger.debug("label no change")
                else:
                    shape.label = result[1][0]
            else:
                logger.warning("Can not recognise the box")
                if self.noLabelText == shape.label:
                    logger.debug("label no change")
                else:
                    shape.label = self.noLabelText
            self.singleLabel(shape)
            self.setDirty()

    def exportJSON(self):
        """
        export PPLabel and CSV to JSON (PubTabNet)
        """

        # automatically save annotations
        self.saveFilestate()
        self.savePPlabel(mode="auto")

        # load box annotations
        labeldict = {}
        if not os.path.exists(self.PPlabelpath):
            msg = "ERROR, Can not find Label.txt"
            QMessageBox.information(self, "Information", msg)
            return
        else:
            with open(self.PPlabelpath, "r", encoding="utf-8") as f:
                data = f.readlines()
                for each in data:
                    file, label = each.split("\t")
                    if label:
                        label = label.replace("false", "False")
                        label = label.replace("true", "True")
                        label = label.replace("null", "None")
                        labeldict[file] = eval(label)
                    else:
                        labeldict[file] = []

        # read table recognition output
        TableRec_excel_dir = os.path.join(self.lastOpenDir, "tableRec_excel_output")

        # save txt
        fid = open("{}/gt.txt".format(self.lastOpenDir), "w", encoding="utf-8")
        for image_path in labeldict.keys():
            # load csv annotations
            filename, _ = os.path.splitext(os.path.basename(image_path))
            csv_path = os.path.join(TableRec_excel_dir, filename + ".xlsx")
            if not os.path.exists(csv_path):
                continue

            excel = openpyxl.load_workbook(csv_path, data_only=True)
            sheet0 = excel.worksheets[0]  # only sheet 0
            merged_cells = sheet0.merged_cells.ranges  # list of merged cell ranges

            html_list = [["td"] * sheet0.max_column for i in range(sheet0.max_row)]

            for merged in merged_cells:
                # Convert merged cell range to start row, end row, start col, end col
                sr = merged.min_row - 1
                er = merged.max_row - 1
                sc = merged.min_col - 1
                ec = merged.max_col - 1
                html_list = expand_list((sr, er, sc, ec), html_list)

            token_list = convert_token(html_list)

            # load box annotations
            cells = []
            for anno in labeldict[image_path]:
                tokens = list(anno["transcription"])
                cells.append({"tokens": tokens, "bbox": anno["points"]})

            # 构造标注信息
            html = {"structure": {"tokens": token_list}, "cells": cells}
            d = {"filename": os.path.basename(image_path), "html": html}
            # 重构HTML
            d["gt"] = rebuild_html_from_ppstructure_label(d)
            fid.write("{}\n".format(json.dumps(d, ensure_ascii=False)))

        # convert to PP-Structure label format
        fid.close()
        msg = "JSON successfully saved in {}/gt.txt".format(self.lastOpenDir)
        QMessageBox.information(self, "Information", msg)

    def autolcm(self):
        vbox = QVBoxLayout()
        hbox = QHBoxLayout()
        self.panel = QLabel()
        self.panel.setText(self.stringBundle.getString("choseModelLg"))
        self.panel.setAlignment(Qt.AlignLeft)
        self.comboBox = QComboBox()
        self.comboBox.setObjectName("comboBox")
        self.comboBox.addItems(
            ["Chinese & English", "English", "French", "German", "Korean", "Japanese"]
        )
        vbox.addWidget(self.panel)
        vbox.addWidget(self.comboBox)
        self.dialog = QDialog()
        self.dialog.resize(300, 100)
        self.okBtn = QPushButton(self.stringBundle.getString("ok"))
        self.cancelBtn = QPushButton(self.stringBundle.getString("cancel"))

        self.okBtn.clicked.connect(self.modelChoose)
        self.cancelBtn.clicked.connect(self.cancel)
        self.dialog.setWindowTitle(self.stringBundle.getString("choseModelLg"))

        hbox.addWidget(self.okBtn)
        hbox.addWidget(self.cancelBtn)

        vbox.addWidget(self.panel)
        vbox.addLayout(hbox)
        self.dialog.setLayout(vbox)
        self.dialog.setWindowModality(Qt.ApplicationModal)
        self.dialog.exec_()
        if self.filePath:
            self.AutoRecognition.setEnabled(True)
            self.actions.AutoRec.setEnabled(True)

    def modelChoose(self):
        current_text = self.comboBox.currentText()
        logger.debug("Model selected: %s", current_text)
        lg_idx = {
            "Chinese & English": "ch",
            "English": "en",
            "French": "french",
            "German": "german",
            "Korean": "korean",
            "Japanese": "japan",
        }
        if current_text in lg_idx:
            choose_lang = lg_idx[current_text]
            if hasattr(self, "ocr"):
                del self.ocr
                self.ocr = PaddleOCR(
                    use_doc_orientation_classify=False,
                    use_textline_orientation=False,
                    use_doc_unwarping=False,
                    lang=choose_lang,
                    device=self.gpu,
                )
            if choose_lang in ["ch", "en"]:
                if hasattr(self, "table_ocr"):
                    del self.table_ocr
                self.table_ocr = PPStructureV3(
                    use_doc_orientation_classify=False,
                    use_doc_unwarping=False,
                    use_seal_recognition=False,
                    use_table_recognition=True,
                    use_formula_recognition=False,
                    use_chart_recognition=False,
                    use_region_detection=False,
                    device=self.gpu,
                )
        else:
            logger.error("Invalid language selection")
        self.dialog.close()

    def cancel(self):
        self.dialog.close()

    def loadFilestate(self, saveDir):
        self.fileStatepath = saveDir + "/fileState.txt"
        self.fileStatedict = {}
        if not os.path.exists(self.fileStatepath):
            f = open(self.fileStatepath, "w", encoding="utf-8")
        else:
            with open(self.fileStatepath, "r", encoding="utf-8") as f:
                states = f.readlines()
                for each in states:
                    file, state = each.split("\t")
                    self.fileStatedict[self.getImglabelidx(file)] = 1
                self.actions.saveLabel.setEnabled(True)
                self.actions.saveRec.setEnabled(True)
                self.actions.exportJSON.setEnabled(True)

    def saveFilestate(self):
        with open(self.fileStatepath, "w", encoding="utf-8") as f:
            for key in self.fileStatedict:
                f.write(key + "\t")
                f.write(str(self.fileStatedict[key]) + "\n")

    def loadLabelFile(self, labelpath):
        labeldict = {}
        if not os.path.exists(labelpath):
            f = open(labelpath, "w", encoding="utf-8")

        else:
            with open(labelpath, "r", encoding="utf-8") as f:
                data = f.readlines()
                for each in data:
                    file, label = each.split("\t")
                    if label:
                        label = label.replace("false", "False")
                        label = label.replace("true", "True")
                        label = label.replace("null", "None")
                        labeldict[file] = eval(label)
                    else:
                        labeldict[file] = []
        return labeldict

    def savePPlabel(self, mode="Manual"):
        savedfile = [self.getImglabelidx(i) for i in self.fileStatedict.keys()]
        with open(self.PPlabelpath, "w", encoding="utf-8") as f:
            for key in self.PPlabel:
                if key in savedfile and self.PPlabel[key] != []:
                    f.write(key + "\t")
                    f.write(json.dumps(self.PPlabel[key], ensure_ascii=False) + "\n")

        if mode == "Manual":
            if self.lang == "ch":
                msg = "已将检查过的图片标签保存在 " + self.PPlabelpath + " 文件中"
            else:
                msg = "Images that have been checked are saved in " + self.PPlabelpath
            QMessageBox.information(self, "Information", msg)

    def saveCacheLabel(self):
        with open(self.Cachelabelpath, "w", encoding="utf-8") as f:
            for key in self.Cachelabel:
                f.write(key + "\t")
                f.write(json.dumps(self.Cachelabel[key], ensure_ascii=False) + "\n")

    def saveLabelFile(self):
        self.saveFilestate()
        self.savePPlabel()

    def saveRecResult(self):
        if {} in [self.PPlabelpath, self.PPlabel, self.fileStatedict]:
            QMessageBox.information(self, "Information", "Check the image first")
            return

        base_dir = os.path.dirname(self.PPlabelpath)
        rec_gt_dir = base_dir + "/rec_gt.txt"
        crop_img_dir = base_dir + "/crop_img/"
        ques_img = []
        if not os.path.exists(crop_img_dir):
            os.mkdir(crop_img_dir)

        with open(rec_gt_dir, "w", encoding="utf-8") as f:
            for key in self.fileStatedict:
                idx = self.getImglabelidx(key)
                try:
                    img_path = os.path.dirname(base_dir) + "/" + key
                    img = cv2.imdecode(np.fromfile(img_path, dtype=np.uint8), -1)
                    for i, label in enumerate(self.PPlabel[idx]):
                        if label["difficult"]:
                            continue
                        img_crop = get_rotate_crop_image(
                            img, np.array(label["points"], np.float32)
                        )
                        img_name = (
                            os.path.splitext(os.path.basename(idx))[0]
                            + "_crop_"
                            + str(i)
                            + ".jpg"
                        )
                        cv2.imencode(".jpg", img_crop)[1].tofile(
                            crop_img_dir + img_name
                        )
                        f.write("crop_img/" + img_name + "\t")
                        f.write(label["transcription"] + "\n")
                except KeyError as e:
                    pass
                except Exception as e:
                    ques_img.append(key)
                    logger.exception("Error processing image %s: %s", key, e)
        if ques_img:
            QMessageBox.information(
                self,
                "Information",
                "The following images can not be saved, please check the image path and labels.\n"
                + "".join(str(i) + "\n" for i in ques_img),
            )
        QMessageBox.information(
            self,
            "Information",
            "Cropped images have been saved in " + str(crop_img_dir),
        )

    def speedChoose(self):
        if self.labelDialogOption.isChecked():
            self.canvas.newShape.disconnect()
            self.canvas.newShape.connect(partial(self.newShape, True))

        else:
            self.canvas.newShape.disconnect()
            self.canvas.newShape.connect(partial(self.newShape, False))

    def autoSaveFunc(self):
        if self.autoSaveOption.isChecked():
            self.autoSaveNum = 1  # Real auto_Save
            try:
                self.saveLabelFile()
            except Exception:
                pass
            logger.info(
                "The program will automatically save once after confirming an image"
            )
        else:
            self.autoSaveNum = 5  # Used for backup
            logger.info(
                "The program will automatically save once after confirming 5 images (default)"
            )

    def change_box_key(self):
        if not self.kie_mode:
            return
        key_text, _ = self.keyDialog.popUp(self.key_previous_text)
        if key_text is None:
            return
        self.key_previous_text = key_text
        for shape in self.canvas.selectedShapes:
            shape.key_cls = key_text
            if not self.keyList.findItemsByLabel(key_text):
                item = self.keyList.createItemFromLabel(key_text)
                self.keyList.addItem(item)
                rgb = self._get_rgb_by_label(key_text, self.kie_mode)
                self.keyList.setItemLabel(item, key_text, rgb)

            self._update_shape_color(shape)
            self.keyDialog.addLabelHistory(key_text)

        # save changed shape
        self.setDirty()

    def undoShapeEdit(self):
        self.canvas.restoreShape()
        self.labelList.clear()
        self.indexList.clear()
        self.BoxList.clear()
        self.loadShapes(self.canvas.shapes)
        self.actions.undo.setEnabled(self.canvas.isShapeRestorable)

    def loadShapes(self, shapes, replace=True):
        self._noSelectionSlot = True
        for shape in shapes:
            self.addLabel(shape)
        self.labelList.clearSelection()
        self.indexList.clearSelection()
        self._noSelectionSlot = False
        self.canvas.loadShapes(shapes, replace=replace)
        logger.debug("loadShapes")

    def lockSelectedShape(self):
        """lock the selected shapes.

        Add self.selectedShapes to lock self.canvas.lockedShapes,
        which holds the ratio of the four coordinates of the locked shapes
        to the width and height of the image
        """
        width, height = self.image.width(), self.image.height()

        def format_shape(s):
            return dict(
                label=s.label,  # str
                line_color=s.line_color.getRgb(),
                fill_color=s.fill_color.getRgb(),
                ratio=[
                    [int(p.x()) / width, int(p.y()) / height] for p in s.points
                ],  # QPonitF
                difficult=s.difficult,  # bool
                key_cls=s.key_cls,  # bool
            )

        # lock
        if len(self.canvas.lockedShapes) == 0:
            for s in self.canvas.selectedShapes:
                s.line_color = DEFAULT_LOCK_COLOR
                s.locked = True
            shapes = [format_shape(shape) for shape in self.canvas.selectedShapes]
            trans_dic = []
            for box in shapes:
                trans_dict = {
                    "transcription": box["label"],
                    "ratio": box["ratio"],
                    "difficult": box["difficult"],
                }
                if self.kie_mode:
                    trans_dict.update({"key_cls": box["key_cls"]})
                trans_dic.append(trans_dict)
            self.canvas.lockedShapes = trans_dic
            self.actions.save.setEnabled(True)

        # unlock
        else:
            for s in self.canvas.shapes:
                s.line_color = DEFAULT_LINE_COLOR
            self.canvas.lockedShapes = []
            self.result_dic_locked = []
            self.setDirty()
            self.actions.save.setEnabled(True)

    def expandSelectedShape(self):
        img = cv2.imdecode(np.fromfile(self.filePath, dtype=np.uint8), 1)
        for shape in self.canvas.selectedShapes:
            box = [[int(p.x()), int(p.y())] for p in shape.points]
            if len(box) > 4:
                box = self.gen_quad_from_poly(np.array(box))
            assert len(box) == 4
            box = boxPad(box, img.shape, 3)
            shape.points = [
                QPointF(box[0][0], box[0][1]),
                QPointF(box[1][0], box[1][1]),
                QPointF(box[2][0], box[2][1]),
                QPointF(box[3][0], box[3][1]),
            ]
            logger.debug("Shape points: %s", shape.points)
            self.updateBoxlist()
            self.setDirty()

    def sort_rectangles(self, rectangles, row_height_threshold=0.5):
        if not rectangles:
            return []

        def get_top_left(rect):
            xs = [p[0] for p in rect]
            ys = [p[1] for p in rect]
            return (min(xs), min(ys))

        avg_height = sum(
            [max(p[1] for p in rect) - min(p[1] for p in rect) for rect in rectangles]
        ) / len(rectangles)
        threshold = avg_height * row_height_threshold
        indexed_rects = [(i, get_top_left(rect)) for i, rect in enumerate(rectangles)]
        indexed_rects.sort(key=lambda x: x[1][1])
        rows = []
        current_row = []
        last_y = indexed_rects[0][1][1]
        for item in indexed_rects:
            i, (x, y) = item
            if abs(y - last_y) <= threshold:
                current_row.append(item)
            else:
                rows.append(current_row)
                current_row = [item]
            last_y = y
        if current_row:
            rows.append(current_row)
        sorted_rects = []
        for row in rows:
            row.sort(key=lambda x: x[1][0])
            sorted_rects.extend([rectangles[i] for i, _ in row])
        return sorted_rects

    def resortBoxPosition(self):
        # get original elements
        items = []
        for i in range(self.BoxList.count()):
            item = self.BoxList.item(i)
            items.append({"text": item.text(), "object": item})
        # get coordinate points
        rectangles = []
        for item in items:
            text = item["text"]
            try:
                rect = ast.literal_eval(text)  # 转为列表
                rectangles.append(rect)
            except (ValueError, SyntaxError) as e:
                logger.error(f"Error parsing text: {text}")
                continue
        # start resort
        sorted_rectangles = self.sort_rectangles(rectangles, row_height_threshold=0.5)
        # old_idx <--> new_idx
        index_map = []
        for sorted_rect in sorted_rectangles:
            for old_idx, rect in enumerate(rectangles):
                if rect == sorted_rect:
                    index_map.append(old_idx)
                    break
        # resort BoxList labelList canvas.shapes
        items = [self.BoxList.takeItem(0) for _ in range(self.BoxList.count())]
        items_label = [
            self.labelList.takeItem(0) for _ in range(self.labelList.count())
        ]
        shapes = self.canvas.shapes
        self.canvas.shapes = []
        for new_idx in range(len(index_map)):
            old_idx = index_map[new_idx]
            self.BoxList.insertItem(new_idx, items[old_idx])
            self.labelList.insertItem(new_idx, items_label[old_idx])
            self.canvas.shapes.insert(new_idx, shapes[old_idx])
        QMessageBox.information(
            self,
            "Information",
            "resort success!",
        )

    def importhtml(self):
        if not self.dict_html:
            parent_dir = os.path.dirname(self.lastOpenDir)
            self.htmlfile_path = os.path.join(parent_dir, "val_html.txt")
            tablepyxl.convert_html_txt_to_dict(self.htmlfile_path, self.dict_html)
            excel_dir = os.path.join(parent_dir, "output_excel")
            os.makedirs(excel_dir, exist_ok=True)
            for key, value in self.dict_html.items():
                excel_name = key.rsplit(".", 1)[0] + ".xlsx"
                excel_path = os.path.join(excel_dir, excel_name)
                tablepyxl.html_table_to_excel_complex(value, excel_path)
                self.dict_excel[key] = excel_path
        filePath = self.mImgList[self.currIndex]
        filePath_base = os.path.basename(filePath)

        open_excel_path = self.dict_excel[filePath_base]

        self.dict_export[filePath_base] = open_excel_path
        os.system("open " + os.path.normpath(open_excel_path))

    def exporthtml(self):
        parent_directory = os.path.dirname(self.htmlfile_path)
        new_directory = os.path.join(parent_directory, "backup")
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name_with_extension = os.path.basename(self.htmlfile_path)
        file_name, file_extension = os.path.splitext(file_name_with_extension)
        new_file_name = f"{file_name}_{current_time}{file_extension}"
        new_htmlfile = os.path.join(new_directory, new_file_name)
        if not os.path.exists(new_directory):
            os.makedirs(new_directory)
        tablepyxl.save_dict_to_html_txt(self.dict_html, new_htmlfile)
        for key, value in self.dict_export.items():
            html_content = tablepyxl.xl_to_html(value)
            self.dict_html[key] = html_content
            if self.autoCheck.isChecked():
                parent_dir = os.path.dirname(self.lastOpenDir)
                excel_dir = os.path.join(parent_dir, "check_excel")
                excel_name = key.rsplit(".", 1)[0] + ".xlsx"
                excel_path = os.path.join(excel_dir, excel_name)
                tablepyxl.html_table_to_excel_complex(html_content, excel_path)
                os.system("open " + os.path.normpath(excel_path))
        self.dict_export = {}
        tablepyxl.save_dict_to_html_txt(self.dict_html, self.htmlfile_path)


def inverted(color):
    return QColor(*[255 - v for v in color.getRgb()])


def read(filename, default=None):
    try:
        with open(filename, "rb") as f:
            return f.read()
    except Exception:
        return default


def str2bool(v):
    return v.lower() in ("true", "t", "1")


def parse_rgb(value):
    r, g, b = value.split(",")
    r, g, b = int(r), int(g), int(b)
    if not (0 <= r <= 255 and 0 <= g <= 255 and 0 <= b <= 255):
        raise argparse.ArgumentTypeError("RGB values must be between 0 and 255.")
    return (r, g, b)


def get_main_app(argv=[]):
    """
    Standard boilerplate Qt application code.
    Do everything but app.exec_() -- so that we can test the application in one thread
    """
    app = QApplication(argv)
    app.setApplicationName(__appname__)
    app.setWindowIcon(newIcon("app"))
    # Tzutalin 201705+: Accept extra arguments to change predefined class file
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("--lang", type=str, default="ch", nargs="?")
    arg_parser.add_argument("--gpu", type=str2bool, default=True, nargs="?")
    arg_parser.add_argument(
        "--img_list_natural_sort", type=str2bool, default=True, nargs="?"
    )
    arg_parser.add_argument("--kie", type=str2bool, default=False, nargs="?")
    arg_parser.add_argument(
        "--predefined_classes_file",
        default=os.path.join(
            os.path.dirname(__file__), "data", "predefined_classes.txt"
        ),
        nargs="?",
    )
    arg_parser.add_argument("--det_model_dir", type=str, default=None, nargs="?")
    arg_parser.add_argument("--rec_model_dir", type=str, default=None, nargs="?")
    arg_parser.add_argument("--rec_char_dict_path", type=str, default=None, nargs="?")
    arg_parser.add_argument("--cls_model_dir", type=str, default=None, nargs="?")
    arg_parser.add_argument(
        "--bbox_auto_zoom_center", type=str2bool, default=False, nargs="?"
    )
    arg_parser.add_argument("--label_font_path", type=str, default=None, nargs="?")
    arg_parser.add_argument(
        "--selected_shape_color",
        type=parse_rgb,
        default="255,255,0",
        nargs="?",
        help='An RGB value as "R,G,B".',
    )

    args = arg_parser.parse_args(argv[1:])

    win = MainWindow(
        lang=args.lang,
        gpu=args.gpu,
        img_list_natural_sort=args.img_list_natural_sort,
        kie_mode=args.kie,
        default_predefined_class_file=args.predefined_classes_file,
        det_model_dir=args.det_model_dir,
        rec_model_dir=args.rec_model_dir,
        cls_model_dir=args.cls_model_dir,
        bbox_auto_zoom_center=args.bbox_auto_zoom_center,
        label_font_path=args.label_font_path,
        selected_shape_color=args.selected_shape_color,
    )
    win.show()
    return app, win


def main():
    """construct main app and run it"""
    app, _win = get_main_app(sys.argv)
    return app.exec_()


if __name__ == "__main__":
    sys.exit(main())
